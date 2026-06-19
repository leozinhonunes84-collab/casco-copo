from __future__ import annotations

import os
import re
import time
import unicodedata
from base64 import b64decode
from collections.abc import Callable
from csv import reader as csv_reader
from datetime import datetime
from io import BytesIO, StringIO
from pathlib import Path

import openpyxl
from playwright.sync_api import Page, sync_playwright

from ajuste_precos_dashboard import (
    abrir_menu_produtos,
    bool_env,
    buscar_produto,
    env_required,
    selecionar_local,
)
from importar_fiscal_dashboard import (
    abrir_modal_edicao_produtos_excel,
    abrir_modal_importacao,
    baixar_tabela_exemplo,
    desativar_status_ativo_no_modal,
    importar_arquivo_atual,
    sheet_value,
)
from sheets_prices import load_env_file, money_to_cents


LogFn = Callable[[str], None]

UNIQUE_STEP_LABELS = {
    "register_product": "Cadastrar produto",
    "activate_product": "Liberar produto",
    "adjust_fiscal": "Ajustar fiscal",
}

DEFAULT_UNIQUE_STEPS = ["register_product", "activate_product", "adjust_fiscal"]


def log_default(message: str) -> None:
    print(message)


def normalize_key(value: object) -> str:
    text = str(value or "").strip().upper()
    text = "".join(
        char
        for char in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(char)
    )
    text = text.replace("Ã‡", "C").replace("Ç", "C")
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_sku(value: object) -> str:
    text = str(value or "").strip()
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    return text.removeprefix("#").strip()


def parse_price_to_cents(value: object) -> int:
    if value in (None, "") or not str(value).strip():
        raise ValueError("preco vazio")
    if isinstance(value, (int, float)):
        amount = float(value)
        if amount > 999 and float(amount).is_integer():
            return int(amount)
        return int(round(amount * 100))
    return money_to_cents(value)


def _header_indexes(headers: list[object]) -> dict[str, int | None]:
    indexes: dict[str, int | None] = {
        "sku": 0,
        "price": 1,
        "name": None,
        "ncm": None,
        "cest": None,
        "fiscal_group": None,
    }

    for index, header in enumerate(headers):
        key = normalize_key(header)
        if "SKU" in key or key in {"CODIGO", "CODIGO PRODUTO", "ID"}:
            indexes["sku"] = index
        if "PRECO" in key or key.startswith("PRE") or "VALOR" in key:
            indexes["price"] = index
        if key in {"PRODUTO", "NOME", "ITEM"}:
            indexes["name"] = index
        if key == "NCM" or "FISCAL - NCM" in key:
            indexes["ncm"] = index
        if key == "CEST" or "FISCAL - CEST" in key:
            indexes["cest"] = index
        if "GRUPO FISCAL" in key:
            indexes["fiscal_group"] = index
    return indexes


def _looks_like_header(row: list[object]) -> bool:
    return any(re.search(r"[A-Za-zÀ-ÿ]", str(cell or "")) for cell in row)


def _dedupe_products(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    unique: list[dict[str, object]] = []
    seen: set[str] = set()
    for row in rows:
        sku = str(row["sku"])
        if sku in seen:
            continue
        unique.append(row)
        seen.add(sku)
    return unique


def _value_by_index(row: list[object], index: int | None) -> object:
    if index is None or index >= len(row):
        return ""
    return row[index]


def _index_or(index: int | None, default: int) -> int:
    return default if index is None else int(index)


def normalize_unique_steps(raw_steps: object) -> list[str]:
    if not isinstance(raw_steps, list):
        return DEFAULT_UNIQUE_STEPS.copy()
    steps: list[str] = []
    for item in raw_steps:
        step = str(item)
        if step in UNIQUE_STEP_LABELS and step not in steps:
            steps.append(step)
    return steps or DEFAULT_UNIQUE_STEPS.copy()


def parse_unique_products(raw_rows: object = "", file_info: dict[str, object] | None = None) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []

    if file_info:
        filename = str(file_info.get("name") or "produtos.xlsx").strip()
        content_b64 = str(file_info.get("content_base64") or "")
        if not content_b64:
            raise ValueError("Arquivo vazio")

        data = b64decode(content_b64)
        suffix = Path(filename).suffix.lower()
        if suffix in {".xlsx", ".xlsm"}:
            workbook = openpyxl.load_workbook(BytesIO(data), data_only=True)
            worksheet = workbook.active
            header = [worksheet.cell(1, col).value for col in range(1, worksheet.max_column + 1)]
            indexes = _header_indexes(header)
            start_row = 2 if _looks_like_header(header) else 1
            for row_idx in range(start_row, worksheet.max_row + 1):
                sku_col = _index_or(indexes["sku"], 0)
                price_col = _index_or(indexes["price"], 1)
                sku = normalize_sku(worksheet.cell(row_idx, sku_col + 1).value)
                price_raw = worksheet.cell(row_idx, price_col + 1).value
                if not sku:
                    continue
                name_col = indexes["name"]
                name = worksheet.cell(row_idx, int(name_col) + 1).value if name_col is not None else ""
                ncm_col = indexes["ncm"]
                cest_col = indexes["cest"]
                fiscal_group_col = indexes["fiscal_group"]
                rows.append(
                    {
                        "sku": sku,
                        "produto": str(name or "").strip(),
                        "preco_centavos": parse_price_to_cents(price_raw) if price_raw not in (None, "") else None,
                        "fiscal_ncm": str(worksheet.cell(row_idx, int(ncm_col) + 1).value or "").strip()
                        if ncm_col is not None
                        else "",
                        "fiscal_cest": str(worksheet.cell(row_idx, int(cest_col) + 1).value or "").strip()
                        if cest_col is not None
                        else "",
                        "fiscal_group": str(worksheet.cell(row_idx, int(fiscal_group_col) + 1).value or "").strip()
                        if fiscal_group_col is not None
                        else "",
                    }
                )
        elif suffix == ".csv":
            text = data.decode("utf-8-sig")
            sample = text[:2000]
            delimiter = ";" if sample.count(";") >= sample.count(",") else ","
            csv_rows = list(csv_reader(StringIO(text), delimiter=delimiter))
            if csv_rows:
                indexes = _header_indexes(csv_rows[0])
                sku_index = _index_or(indexes["sku"], 0)
                price_index = _index_or(indexes["price"], 1)
                start = 1 if _looks_like_header(csv_rows[0]) else 0
                for row in csv_rows[start:]:
                    if sku_index >= len(row):
                        continue
                    sku = normalize_sku(row[sku_index])
                    price_raw = row[price_index] if price_index < len(row) else ""
                    if not sku:
                        continue
                    name = _value_by_index(row, indexes["name"])
                    rows.append(
                        {
                            "sku": sku,
                            "produto": str(name or "").strip(),
                            "preco_centavos": parse_price_to_cents(price_raw)
                            if str(price_raw or "").strip()
                            else None,
                            "fiscal_ncm": str(_value_by_index(row, indexes["ncm"]) or "").strip(),
                            "fiscal_cest": str(_value_by_index(row, indexes["cest"]) or "").strip(),
                            "fiscal_group": str(_value_by_index(row, indexes["fiscal_group"]) or "").strip(),
                        }
                    )
        else:
            raise ValueError("Arquivo deve ser .xlsx, .xlsm ou .csv")
    else:
        for raw_line in str(raw_rows or "").splitlines():
            line = raw_line.strip()
            if not line:
                continue
            if ";" in line:
                parts = [part.strip() for part in line.split(";")]
            elif "\t" in line:
                parts = [part.strip() for part in line.split("\t")]
            else:
                parts = line.split()
            if not parts:
                continue
            rows.append(
                {
                    "sku": normalize_sku(parts[0]),
                    "produto": "",
                    "preco_centavos": parse_price_to_cents(parts[1]) if len(parts) >= 2 and parts[1] else None,
                    "fiscal_ncm": str(parts[2]).strip() if len(parts) >= 3 else "",
                    "fiscal_cest": str(parts[3]).strip() if len(parts) >= 4 else "",
                    "fiscal_group": str(parts[4]).strip() if len(parts) >= 5 else "",
                }
            )

    rows = _dedupe_products([row for row in rows if row.get("sku")])
    if not rows:
        raise ValueError("Nenhum produto valido encontrado. Informe pelo menos o SKU.")
    return rows


def parse_table_price(value: object) -> int | None:
    text = str(value or "")
    match = re.search(r"R\$\s*([\d.,]+)", text)
    if not match:
        return None
    try:
        return money_to_cents(match.group(1))
    except Exception:
        return None


def cents_to_brl_text(value: int) -> str:
    reais = int(value) / 100
    return f"{reais:.2f}".replace(".", ",")


def fiscal_pendente(value: object) -> bool:
    text = normalize_key(value)
    return text.startswith("NAO") or text in {"", "-", "N"}


def produto_na_tabela(page: Page, sku: str) -> dict[str, object]:
    return page.evaluate(
        """
        (sku) => {
            const normalize = (value) => String(value || "")
                .normalize("NFD")
                .replace(/[\\u0300-\\u036f]/g, "")
                .replace(/\\s+/g, " ")
                .trim()
                .toUpperCase();
            const wanted = String(sku || "").replace(/^#/, "").trim();
            const headers = [...document.querySelectorAll("table thead th")]
                .map((th) => normalize(th.innerText || th.textContent));
            const indexOf = (needle) => headers.findIndex((header) => header.includes(needle));
            const skuIdx = indexOf("SKU");
            const nameIdx = indexOf("PRODUTO");
            const mountableIdx = indexOf("MONTAVEL");
            const fiscalIdx = indexOf("DADOS FISCAIS");
            const priceIdx = indexOf("PRECO");
            const rows = [...document.querySelectorAll("table tbody tr")];

            for (const row of rows) {
                const cells = [...row.querySelectorAll("td")].map((td) => (td.innerText || td.textContent || "").trim());
                const skuCell = skuIdx >= 0 && skuIdx < cells.length ? cells[skuIdx] : "";
                const anySku = cells.some((cell) => cell.replace(/^#/, "").trim() === wanted);
                if (skuCell.replace(/^#/, "").trim() !== wanted && !anySku) continue;

                return {
                    found: true,
                    sku: wanted,
                    produto: nameIdx >= 0 ? cells[nameIdx] || "" : "",
                    produto_montavel: mountableIdx >= 0 ? cells[mountableIdx] || "" : "",
                    dados_fiscais: fiscalIdx >= 0 ? cells[fiscalIdx] || "" : "",
                    preco_atual_texto: priceIdx >= 0 ? cells[priceIdx] || "" : "",
                };
            }
            return {found: false, sku: wanted};
        }
        """,
        sku,
    )


def coluna_opcional(worksheet, header: str) -> int | None:
    wanted = normalize_key(header)
    for col in range(1, worksheet.max_column + 1):
        if normalize_key(worksheet.cell(1, col).value) == wanted:
            return col
    return None


def set_cell_if_column_exists(worksheet, row_idx: int, header: str, value: object) -> None:
    col = coluna_opcional(worksheet, header)
    if col is not None:
        worksheet.cell(row_idx, col).value = value


def sheet_value_or_text(workbook, sheet_name: str, value: object, default: str) -> object:
    text = str(value or default).strip() or default
    try:
        return sheet_value(workbook, sheet_name, text)
    except Exception:
        if text != default:
            return sheet_value(workbook, sheet_name, default)
        raise


def localizar_linha_sku_planilha(worksheet, sku: str) -> int:
    sku_col = coluna_opcional(worksheet, "SKU")
    wanted = normalize_sku(sku)
    if sku_col is not None:
        for row_idx in range(2, worksheet.max_row + 1):
            if normalize_sku(worksheet.cell(row_idx, sku_col).value) == wanted:
                return row_idx

    data_rows = [
        row_idx
        for row_idx in range(2, worksheet.max_row + 1)
        if any(worksheet.cell(row_idx, col).value not in (None, "") for col in range(1, worksheet.max_column + 1))
    ]
    if len(data_rows) == 1:
        return data_rows[0]
    raise RuntimeError(f"SKU #{sku} nao localizado na planilha baixada da ZigPay")


def fiscal_value_by_header(fiscal_values: dict[str, object], needle: str) -> object:
    wanted = normalize_key(needle)
    for header, value in fiscal_values.items():
        if wanted in normalize_key(header):
            return value
    return ""


def source_value_by_header(source_values: dict[str, object], header: str) -> object:
    return source_values.get(normalize_key(header), "")


def sheet_lookup_or_raw(workbook, sheet_name: str, value: object) -> object:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return sheet_value(workbook, sheet_name, text)
    except RuntimeError:
        return text


def sheet_lookup_or_none(workbook, sheet_name: str, value: object) -> object:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return sheet_value(workbook, sheet_name, text)
    except RuntimeError:
        return None


def extrair_campos_origem_planilha_produto_unico(template_path: Path, sku: str) -> dict[str, object]:
    workbook = openpyxl.load_workbook(template_path, data_only=False)
    worksheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.active
    row_idx = localizar_linha_sku_planilha(worksheet, sku)
    wanted = {
        "NOME *",
        "TIPO DE PRODUTO *",
        "CATEGORIA *",
        "CATEGORIA DO MENU",
        "BARES (SEPARADOS POR PONTO E VIRGULA)",
        "ID DO PRODUTO DE SISTEMA",
        "UNIDADE DE MEDIDA",
        "ESTOCAVEL",
        "CONTEM ALCOOL?",
        "NAO EXIBIR PRODUTO NO APLICATIVO ZIGAPP",
        "DESCRICAO",
        "IMAGEM (ID)",
        "IMAGEM ORIGINAL",
    }
    values: dict[str, object] = {}
    for col in range(1, worksheet.max_column + 1):
        header = normalize_key(worksheet.cell(1, col).value)
        if header in wanted:
            values[header] = worksheet.cell(row_idx, col).value
    return values


def extrair_fiscal_planilha_produto_unico(template_path: Path, sku: str) -> dict[str, object]:
    workbook = openpyxl.load_workbook(template_path, data_only=False)
    worksheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.active
    row_idx = localizar_linha_sku_planilha(worksheet, sku)
    fiscal_values: dict[str, object] = {}
    for col in range(1, worksheet.max_column + 1):
        header = str(worksheet.cell(1, col).value or "").strip()
        if not normalize_key(header).startswith("FISCAL -"):
            continue
        fiscal_values[header] = worksheet.cell(row_idx, col).value

    ncm = str(fiscal_value_by_header(fiscal_values, "NCM") or "").strip()
    if not ncm:
        raise RuntimeError(f"Fiscal da origem esta vazio para SKU #{sku}")
    return fiscal_values


def extrair_preco_planilha_produto_unico(template_path: Path, sku: str) -> int | None:
    workbook = openpyxl.load_workbook(template_path, data_only=True)
    worksheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.active
    row_idx = localizar_linha_sku_planilha(worksheet, sku)

    cents_col = coluna_opcional(worksheet, "PRECO EM CENTAVOS")
    if cents_col is not None:
        cents_value = worksheet.cell(row_idx, cents_col).value
        if cents_value not in (None, ""):
            try:
                return int(round(float(cents_value)))
            except Exception:
                pass

    price_col = coluna_opcional(worksheet, "PRECO *")
    if price_col is not None:
        price_value = worksheet.cell(row_idx, price_col).value
        if price_value not in (None, ""):
            return parse_price_to_cents(price_value)
    return None


def montar_planilha_produto_unico_fiscal(
    template_path: Path,
    unidade: str,
    product: dict[str, object],
    fiscal_values: dict[str, object],
    source_values: dict[str, object],
    fiscal_group: str,
    log: LogFn = log_default,
    create_if_missing: bool = False,
) -> Path:
    workbook = openpyxl.load_workbook(template_path)
    worksheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.active
    sku = str(product["sku"])
    try:
        row_idx = localizar_linha_sku_planilha(worksheet, sku)
    except RuntimeError:
        if not create_if_missing:
            raise
        row_idx = 2
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row_idx, col)
            if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                cell.value = None
    preco_centavos = int(product["preco_centavos"])
    preco_reais = preco_centavos / 100
    grupo = (
        str(fiscal_value_by_header(fiscal_values, "GRUPO FISCAL") or "").strip()
        or str(product.get("fiscal_group") or "").strip()
        or str(fiscal_group or "").strip()
        or "SEM ST COM PIS/COFINS NAO TRIBUTAVEIS"
    )

    set_cell_if_column_exists(worksheet, row_idx, "PRECO *", preco_reais)
    set_cell_if_column_exists(worksheet, row_idx, "PRECO EM CENTAVOS", preco_centavos)
    set_cell_if_column_exists(worksheet, row_idx, "SKU", sku)

    source_name = source_value_by_header(source_values, "NOME *") or product.get("produto", "") or sku
    set_cell_if_column_exists(worksheet, row_idx, "NOME *", source_name)

    source_kind = source_value_by_header(source_values, "TIPO DE PRODUTO *")
    source_category = source_value_by_header(source_values, "CATEGORIA *")
    source_menu_category = source_value_by_header(source_values, "CATEGORIA DO MENU")
    source_bares = source_value_by_header(source_values, "BARES (SEPARADOS POR PONTO E VIRGULA)")
    if source_kind:
        set_cell_if_column_exists(worksheet, row_idx, "TIPO DE PRODUTO *", sheet_lookup_or_raw(workbook, "Kinds", source_kind))
    if source_category:
        set_cell_if_column_exists(worksheet, row_idx, "CATEGORIA *", sheet_lookup_or_raw(workbook, "Categories", source_category))
    if source_menu_category:
        set_cell_if_column_exists(
            worksheet,
            row_idx,
            "CATEGORIA DO MENU",
            sheet_lookup_or_raw(workbook, "Categories", source_menu_category),
        )
    if source_bares:
        bares_value = sheet_lookup_or_none(workbook, "Bares", source_bares)
        if bares_value is not None:
            set_cell_if_column_exists(
                worksheet,
                row_idx,
                "BARES (SEPARADOS POR PONTO E VIRGULA)",
                bares_value,
            )
        else:
            log(f"Campo Bares ignorado em {unidade}: valor '{source_bares}' nao existe no modelo ZigPay.")
    source_system_product = source_value_by_header(source_values, "ID DO PRODUTO DE SISTEMA")
    if source_system_product:
        set_cell_if_column_exists(worksheet, row_idx, "ID DO PRODUTO DE SISTEMA", source_system_product)
    source_description = source_value_by_header(source_values, "DESCRICAO")
    if source_description:
        set_cell_if_column_exists(worksheet, row_idx, "DESCRICAO", source_description)

    for header, value in fiscal_values.items():
        key = normalize_key(header)
        if not key.startswith("FISCAL -") or "PERFIL FISCAL" in key or "GRUPO FISCAL" in key:
            continue
        set_cell_if_column_exists(worksheet, row_idx, header, value)

    set_cell_if_column_exists(
        worksheet,
        row_idx,
        "FISCAL - GRUPO FISCAL *",
        sheet_value_or_text(workbook, "FiscalProductGroups", grupo, "SEM ST COM PIS/COFINS NAO TRIBUTAVEIS"),
    )
    set_cell_if_column_exists(
        worksheet,
        row_idx,
        "FISCAL - PERFIL FISCAL *",
        sheet_value_or_text(workbook, "FiscalProfiles", unidade, unidade),
    )

    unit_value = source_value_by_header(source_values, "UNIDADE DE MEDIDA") or "Unidades"
    stockable_value = source_value_by_header(source_values, "ESTOCAVEL") or "NAO"
    alcohol_value = source_value_by_header(source_values, "CONTEM ALCOOL?") or "NAO"
    hidden_value = source_value_by_header(source_values, "NAO EXIBIR PRODUTO NO APLICATIVO ZIGAPP") or "NAO"
    set_cell_if_column_exists(worksheet, row_idx, "UNIDADE DE MEDIDA", sheet_lookup_or_raw(workbook, "Units", unit_value))
    set_cell_if_column_exists(worksheet, row_idx, "ESTOCAVEL", sheet_lookup_or_raw(workbook, "Boolean", stockable_value))
    set_cell_if_column_exists(worksheet, row_idx, "CONTEM ALCOOL?", sheet_lookup_or_raw(workbook, "Boolean", alcohol_value))
    set_cell_if_column_exists(
        worksheet,
        row_idx,
        "NAO EXIBIR PRODUTO NO APLICATIVO ZIGAPP",
        sheet_lookup_or_raw(workbook, "Boolean", hidden_value),
    )
    if create_if_missing:
        set_cell_if_column_exists(worksheet, row_idx, "ID", None)

    destino = template_path.parent / f"IMPORTAR_PRODUTO_UNICO_{normalize_key(unidade).replace(' ', '_')}_{sku}.xlsx"
    workbook.save(destino)
    log(f"Planilha fiscal do produto unico pronta: {destino.name}")
    return destino


def baixar_planilha_produto_unico(
    page: Page,
    dashboard_url: str,
    unidade: str,
    sku: str,
    log: LogFn = log_default,
    preparar_importacao: bool = False,
) -> tuple[Path, dict[str, object]]:
    page.goto(f"{dashboard_url}/")
    page.wait_for_timeout(2500)
    selecionar_local(page, unidade, log)
    abrir_menu_produtos(page, log)
    buscar_produto(page, sku, log)
    table_info = produto_na_tabela(page, sku)
    if not table_info.get("found"):
        raise RuntimeError(f"SKU #{sku} nao encontrado em {unidade}")

    selecionar_produto_unico_na_tabela(page, sku, log)
    abrir_modal_edicao_produtos_excel(page, log)
    if preparar_importacao:
        desativar_status_ativo_no_modal(page, log)
    template_path = baixar_tabela_exemplo(page, unidade, log)
    if not preparar_importacao:
        try:
            page.keyboard.press("Escape")
            page.wait_for_timeout(800)
        except Exception:
            pass
    return template_path, table_info


def origem_produto_unico(
    page: Page,
    dashboard_url: str,
    fiscal_source: str,
    sku: str,
    cache: dict[str, dict[str, object]],
    log: LogFn = log_default,
) -> dict[str, object]:
    if sku in cache:
        return cache[sku]

    log(f"Buscando preco e fiscal de origem em {fiscal_source} para SKU #{sku}...")
    template_path, table_info = baixar_planilha_produto_unico(
        page,
        dashboard_url,
        fiscal_source,
        sku,
        log,
        preparar_importacao=False,
    )

    preco_centavos = parse_table_price(table_info.get("preco_atual_texto"))
    if preco_centavos is None:
        preco_centavos = extrair_preco_planilha_produto_unico(template_path, sku)

    fiscal_values: dict[str, object] = {}
    fiscal_error = ""
    if fiscal_pendente(table_info.get("dados_fiscais", "")):
        log(
            f"Origem {fiscal_source} aparece como Dados Fiscais = Nao na tela; "
            "usando as colunas fiscais do Excel baixado."
        )
    source_values = extrair_campos_origem_planilha_produto_unico(template_path, sku)
    log(
        "Campos copiados da planilha origem: "
        f"tipo={source_value_by_header(source_values, 'TIPO DE PRODUTO *') or '-'}, "
        f"categoria={source_value_by_header(source_values, 'CATEGORIA *') or '-'}, "
        f"categoria_menu={source_value_by_header(source_values, 'CATEGORIA DO MENU') or '-'}, "
        f"bares={source_value_by_header(source_values, 'BARES (SEPARADOS POR PONTO E VIRGULA)') or '-'}"
    )
    try:
        fiscal_values = extrair_fiscal_planilha_produto_unico(template_path, sku)
    except Exception as exc:
        fiscal_error = str(exc)

    source_info = {
        "preco_centavos": preco_centavos,
        "fiscal_values": fiscal_values,
        "source_values": source_values,
        "fiscal_error": fiscal_error,
        "produto": table_info.get("produto", ""),
    }
    cache[sku] = source_info
    return source_info


def selecionar_produto_unico_na_tabela(page: Page, sku: str, log: LogFn = log_default) -> None:
    def estado_checkbox() -> dict[str, object]:
        return page.evaluate(
            """
            (sku) => {
                const wanted = String(sku || "").replace(/^#/, "").trim();
                const visible = (el) => !!(
                    el
                    && (el.offsetWidth || el.offsetHeight || el.getClientRects().length)
                    && getComputedStyle(el).visibility !== "hidden"
                    && getComputedStyle(el).display !== "none"
                );
                const checked = (el) => {
                    const input = el.matches("input") ? el : el.querySelector("input[type='checkbox']");
                    if (input) return !!input.checked;
                    if (el.getAttribute("aria-checked") === "true") return true;
                    if (el.getAttribute("aria-checked") === "false") return false;
                    return String(el.className).includes("checked");
                };
                const rows = [...document.querySelectorAll("table tbody tr")].filter(visible);
                for (const row of rows) {
                    const cells = [...row.querySelectorAll("td")].map((cell) => (cell.innerText || cell.textContent || "").trim().replace(/^#/, ""));
                    if (!cells.some((cell) => cell === wanted)) continue;
                    const checkbox = [
                        ...row.querySelectorAll("input[type='checkbox'], [role='checkbox'], .ant-checkbox, label")
                    ].filter(visible)[0];
                    if (!checkbox) return {ok: false, reason: `Checkbox nao encontrado para SKU #${wanted}`};
                    const target = checkbox.closest?.("label, [role='checkbox'], .ant-checkbox-wrapper, .ant-checkbox") || checkbox;
                    const box = target.getBoundingClientRect();
                    return {
                        ok: true,
                        checked: checked(target),
                        x: box.left + box.width / 2,
                        y: box.top + box.height / 2,
                    };
                }
                return {ok: false, reason: `Linha nao encontrada para SKU #${wanted}`};
            }
            """,
            sku,
        )

    state = estado_checkbox()
    if not state.get("ok"):
        raise RuntimeError(str(state.get("reason") or f"Nao foi possivel marcar SKU #{sku}"))

    for _ in range(3):
        if state.get("checked"):
            log(f"Produto SKU #{sku} marcado para edicao via Excel")
            return
        page.mouse.click(float(state["x"]), float(state["y"]))
        page.wait_for_timeout(500)
        state = estado_checkbox()

    raise RuntimeError(f"Produto SKU #{sku} nao ficou marcado para edicao via Excel")


def subir_fiscal_produto_unico(
    page: Page,
    dashboard_url: str,
    fiscal_source: str,
    unidade: str,
    product: dict[str, object],
    fiscal_cache: dict[str, dict[str, object]],
    fiscal_group: str,
    log: LogFn = log_default,
) -> None:
    sku = str(product["sku"])
    log(f"Replicando fiscal do SKU #{sku}: {fiscal_source} -> {unidade}")
    source_info = origem_produto_unico(page, dashboard_url, fiscal_source, sku, fiscal_cache, log)
    if source_info.get("fiscal_error"):
        raise RuntimeError(str(source_info["fiscal_error"]))
    fiscal_values = source_info.get("fiscal_values") or {}
    source_values = source_info.get("source_values") or {}
    if not fiscal_values:
        raise RuntimeError(f"Fiscal da origem vazio para SKU #{sku}")
    template_path, _table_info = baixar_planilha_produto_unico(
        page,
        dashboard_url,
        unidade,
        sku,
        log,
        preparar_importacao=True,
    )
    arquivo_importacao = montar_planilha_produto_unico_fiscal(
        template_path,
        unidade,
        product,
        fiscal_values,
        source_values,
        fiscal_group,
        log,
    )
    importar_arquivo_atual(page, arquivo_importacao, log)


def cadastrar_produto_unico(
    page: Page,
    dashboard_url: str,
    fiscal_source: str,
    unidade: str,
    product: dict[str, object],
    fiscal_cache: dict[str, dict[str, object]],
    fiscal_group: str,
    log: LogFn = log_default,
) -> None:
    sku = str(product["sku"])
    log(f"Cadastrando produto unico SKU #{sku} em {unidade} via Excel...")
    source_info = origem_produto_unico(page, dashboard_url, fiscal_source, sku, fiscal_cache, log)
    if source_info.get("fiscal_error"):
        raise RuntimeError(str(source_info["fiscal_error"]))
    fiscal_values = source_info.get("fiscal_values") or {}
    source_values = source_info.get("source_values") or {}
    if not fiscal_values:
        raise RuntimeError(f"Fiscal da origem vazio para cadastrar SKU #{sku}")

    page.goto(f"{dashboard_url}/")
    page.wait_for_timeout(2500)
    selecionar_local(page, unidade, log)
    abrir_menu_produtos(page, log)
    abrir_modal_importacao(page, log)
    desativar_status_ativo_no_modal(page, log)
    template_path = baixar_tabela_exemplo(page, unidade, log)
    arquivo_importacao = montar_planilha_produto_unico_fiscal(
        template_path,
        unidade,
        product,
        fiscal_values,
        source_values,
        fiscal_group,
        log,
        create_if_missing=True,
    )
    importar_arquivo_atual(page, arquivo_importacao, log)


def ativar_produto_unico_na_tabela(page: Page, sku: str, log: LogFn = log_default) -> None:
    def confirmar() -> None:
        try:
            button = page.get_by_role("button", name="OK").first
            if button.is_visible(timeout=1000):
                button.click(timeout=1500)
                page.wait_for_timeout(600)
        except Exception:
            pass

    def estado_switch() -> dict[str, object]:
        return page.evaluate(
            """
            (sku) => {
                const wanted = String(sku || "").replace(/^#/, "").trim();
                const visible = (el) => !!(
                    el
                    && (el.offsetWidth || el.offsetHeight || el.getClientRects().length)
                    && getComputedStyle(el).visibility !== "hidden"
                    && getComputedStyle(el).display !== "none"
                );
                const checked = (el) => {
                    const input = el.matches("input") ? el : el.querySelector("input");
                    if (input) return !!input.checked;
                    if (el.getAttribute("aria-checked") === "true") return true;
                    if (el.getAttribute("aria-checked") === "false") return false;
                    return String(el.className).includes("checked");
                };
                const rows = [...document.querySelectorAll("table tbody tr")].filter(visible);
                for (const row of rows) {
                    const cells = [...row.querySelectorAll("td")].map((cell) => (cell.innerText || cell.textContent || "").trim().replace(/^#/, ""));
                    if (!cells.some((cell) => cell === wanted)) continue;
                    const switches = [...row.querySelectorAll("button[role='switch'], .ant-switch, [role='switch']")]
                        .filter(visible);
                    if (!switches.length) return {ok: false, reason: `Switch Ativado nao encontrado para SKU #${wanted}`};
                    const target = switches.sort((a, b) => {
                        const ar = a.getBoundingClientRect();
                        const br = b.getBoundingClientRect();
                        return br.left - ar.left;
                    })[0];
                    const rect = target.getBoundingClientRect();
                    return {ok: true, checked: checked(target), x: rect.left + rect.width / 2, y: rect.top + rect.height / 2};
                }
                return {ok: false, reason: `SKU #${wanted} nao encontrado para liberar`};
            }
            """,
            sku,
        )

    state = estado_switch()
    if not state.get("ok"):
        raise RuntimeError(str(state.get("reason") or f"Nao foi possivel liberar SKU #{sku}"))
    for _ in range(3):
        if state.get("checked"):
            log(f"Produto SKU #{sku} ja estava liberado/ativo.")
            return
        page.mouse.click(float(state["x"]), float(state["y"]))
        page.wait_for_timeout(900)
        confirmar()
        state = estado_switch()

    if not state.get("checked"):
        raise RuntimeError(f"Produto SKU #{sku} nao ficou ativo apos liberar")
    log(f"Produto SKU #{sku} liberado/ativado.")


def localizar_preco_inline(page: Page, sku: str) -> dict[str, object]:
    return page.evaluate(
        """
        (sku) => {
            const normalize = (value) => String(value || "")
                .normalize("NFD")
                .replace(/[\\u0300-\\u036f]/g, "")
                .replace(/\\s+/g, " ")
                .trim()
                .toUpperCase();
            const visible = (el) => !!(
                el
                && (el.offsetWidth || el.offsetHeight || el.getClientRects().length)
                && getComputedStyle(el).visibility !== "hidden"
                && getComputedStyle(el).display !== "none"
            );
            const wanted = String(sku || "").replace(/^#/, "").trim();
            const headers = [...document.querySelectorAll("table thead th")]
                .map((th) => normalize(th.innerText || th.textContent));
            const priceIdx = headers.findIndex((header) => header.includes("PRECO"));
            const rows = [...document.querySelectorAll("table tbody tr")].filter(visible);
            const row = rows.find((candidate) => [...candidate.querySelectorAll("td")]
                .some((cell) => (cell.innerText || cell.textContent || "").replace(/^#/, "").trim() === wanted));
            if (!row) return {ok: false, reason: `Linha nao encontrada para SKU #${wanted}`};
            const cells = [...row.querySelectorAll("td")];
            const priceCell = priceIdx >= 0 && priceIdx < cells.length
                ? cells[priceIdx]
                : cells.find((cell) => /R\\$|\\d+[,.]\\d{2}/.test(cell.innerText || cell.textContent || ""));
            if (!priceCell) return {ok: false, reason: `Celula de preco nao encontrada para SKU #${wanted}`};
            const rect = priceCell.getBoundingClientRect();
            return {
                ok: true,
                cellX: rect.left + rect.width / 2,
                cellY: rect.top + rect.height / 2,
                cellLeft: rect.left,
                cellRight: rect.right,
                cellTop: rect.top,
                cellBottom: rect.bottom,
            };
        }
        """,
        sku,
    )


def estado_editor_preco_inline(page: Page, sku: str) -> dict[str, object]:
    return page.evaluate(
        """
        (sku) => {
            const normalize = (value) => String(value || "")
                .normalize("NFD")
                .replace(/[\\u0300-\\u036f]/g, "")
                .replace(/\\s+/g, " ")
                .trim()
                .toUpperCase();
            const visible = (el) => !!(
                el
                && (el.offsetWidth || el.offsetHeight || el.getClientRects().length)
                && getComputedStyle(el).visibility !== "hidden"
                && getComputedStyle(el).display !== "none"
            );
            const wanted = String(sku || "").replace(/^#/, "").trim();
            const headers = [...document.querySelectorAll("table thead th")]
                .map((th) => normalize(th.innerText || th.textContent));
            const priceIdx = headers.findIndex((header) => header.includes("PRECO"));
            const rows = [...document.querySelectorAll("table tbody tr")].filter(visible);
            const row = rows.find((candidate) => [...candidate.querySelectorAll("td")]
                .some((cell) => (cell.innerText || cell.textContent || "").replace(/^#/, "").trim() === wanted));
            if (!row) return {ok: false, reason: `Linha nao encontrada para SKU #${wanted}`};
            const cells = [...row.querySelectorAll("td")];
            const priceCell = priceIdx >= 0 && priceIdx < cells.length
                ? cells[priceIdx]
                : cells.find((cell) => /R\\$|\\d+[,.]\\d{2}/.test(cell.innerText || cell.textContent || ""));
            if (!priceCell) return {ok: false, reason: `Celula de preco nao encontrada para SKU #${wanted}`};
            const input = [...priceCell.querySelectorAll("input")]
                .find((candidate) => visible(candidate) && candidate.type !== "checkbox" && !candidate.disabled);
            if (!input) return {ok: true, editing: false};
            const inputRect = input.getBoundingClientRect();
            const buttons = [...priceCell.querySelectorAll("button, [role='button'], svg")]
                .filter(visible)
                .map((el) => {
                    const rect = el.getBoundingClientRect();
                    return {x: rect.left + rect.width / 2, y: rect.top + rect.height / 2, left: rect.left};
                })
                .filter((item) => item.left > inputRect.right - 5)
                .sort((a, b) => a.left - b.left);
            return {
                ok: true,
                editing: true,
                inputX: inputRect.left + inputRect.width / 2,
                inputY: inputRect.top + inputRect.height / 2,
                confirmX: buttons.length ? buttons[buttons.length - 1].x : null,
                confirmY: buttons.length ? buttons[buttons.length - 1].y : null,
            };
        }
        """,
        sku,
    )


def ativar_edicao_preco_inline(page: Page, sku: str, log: LogFn = log_default) -> None:
    state = localizar_preco_inline(page, sku)
    if not state.get("ok"):
        raise RuntimeError(str(state.get("reason") or "Preco nao localizado na tabela"))

    for attempt in range(1, 5):
        editor = estado_editor_preco_inline(page, sku)
        if editor.get("ok") and editor.get("editing"):
            return

        if attempt == 2:
            try:
                button = page.get_by_role("button", name=re.compile("Editar Produtos", re.I)).first
                button.click(timeout=2500)
                page.wait_for_timeout(700)
            except Exception:
                pass

        page.mouse.click(float(state["cellX"]), float(state["cellY"]))
        page.wait_for_timeout(500)
        page.mouse.dblclick(float(state["cellX"]), float(state["cellY"]))
        page.wait_for_timeout(700)

    editor = estado_editor_preco_inline(page, sku)
    if not (editor.get("ok") and editor.get("editing")):
        log(f"Estado editor inline: {editor}")
        raise RuntimeError(f"Campo inline de preco nao abriu para SKU #{sku}")


def alterar_preco_inline_na_tabela(page: Page, sku: str, preco_centavos: int, log: LogFn = log_default) -> None:
    valor = cents_to_brl_text(preco_centavos)
    log(f"Editando preco na tabela -> {valor}")
    ativar_edicao_preco_inline(page, sku, log)
    editor = estado_editor_preco_inline(page, sku)
    if not (editor.get("ok") and editor.get("editing")):
        raise RuntimeError(f"Campo inline de preco nao ficou editavel para SKU #{sku}")

    page.mouse.click(float(editor["inputX"]), float(editor["inputY"]))
    page.wait_for_timeout(200)
    page.keyboard.press("Control+A")
    page.wait_for_timeout(100)
    page.keyboard.type(valor)
    page.wait_for_timeout(300)
    if editor.get("confirmX") is not None and editor.get("confirmY") is not None:
        page.mouse.click(float(editor["confirmX"]), float(editor["confirmY"]))
    else:
        page.keyboard.press("Enter")
    page.wait_for_timeout(2500)


def salvar_relatorio_produto_unico(rows: list[dict[str, object]]) -> Path:
    destino = Path(__file__).resolve().parent / "EXPORTACOES_ZIGPAY"
    destino.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    caminho = destino / f"RELATORIO_PRODUTO_UNICO_{timestamp}.xlsx"

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Resultado"
    headers = [
        "unidade",
        "sku",
        "produto",
        "preco_planilha_centavos",
        "preco_origem_centavos",
        "preco_atual_centavos",
        "fiscal_atual",
        "fiscal_status",
        "fiscal_origem",
        "status",
        "mensagem",
    ]
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])
    for column in worksheet.columns:
        width = max(len(str(cell.value or "")) for cell in column) + 2
        worksheet.column_dimensions[column[0].column_letter].width = min(max(width, 12), 80)
    workbook.save(caminho)
    return caminho


def run_produto_unico_dashboard(
    locais: list[str],
    produtos: object = "",
    log: LogFn = log_default,
    file_info: dict[str, object] | None = None,
    fiscal_source: str = "",
    fiscal_group: str = "SEM ST COM PIS/COFINS NAO TRIBUTAVEIS",
    unique_steps: object = None,
) -> dict[str, object]:
    load_env_file()
    org = env_required("ZIG_ORG")
    user = env_required("ZIG_USER")
    password = env_required("ZIG_PASSWORD")
    dashboard_url = os.environ.get("DASHBOARD_URL", "https://dashboard.zigpay.com.br").rstrip("/")

    locais = [str(local).strip().upper() for local in locais if str(local).strip()]
    fiscal_source = str(fiscal_source or "").strip().upper()
    fiscal_group = str(fiscal_group or "").strip() or "SEM ST COM PIS/COFINS NAO TRIBUTAVEIS"
    steps = normalize_unique_steps(unique_steps)
    product_rows = parse_unique_products(produtos, file_info)
    if not locais:
        raise ValueError("Selecione pelo menos uma unidade")

    rows: list[dict[str, object]] = []
    fiscal_cache: dict[str, dict[str, object]] = {}
    source_errors: dict[str, str] = {}

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(
            headless=bool_env("HEADLESS", default=False),
            slow_mo=60,
        )
        page = browser.new_page(viewport={"width": 1440, "height": 900})

        log("Login no Dashboard ZigPay...")
        page.goto(f"{dashboard_url}/login")
        page.wait_for_selector("input[name='organization']", timeout=15000)
        page.fill("input[name='organization']", org)
        page.fill("input[name='username']", user)
        page.fill("input[name='password']", password)
        page.click("button[type='submit']")
        try:
            page.wait_for_url(lambda url: "/login" not in url, timeout=20000)
        except Exception:
            raise RuntimeError(
                "Login falhou: verifique usuario/senha/organizacao no arquivo .env "
                "(ZIG_USER, ZIG_PASSWORD, ZIG_ORG)"
            )

        if fiscal_source:
            for product in product_rows:
                sku = str(product["sku"])
                try:
                    origem_produto_unico(page, dashboard_url, fiscal_source, sku, fiscal_cache, log)
                except Exception as exc:
                    source_errors[sku] = str(exc)
                    log(f"  ERRO ORIGEM -> SKU #{sku}: {source_errors[sku]}")

        log("Fases Produto Unico: " + " > ".join(UNIQUE_STEP_LABELS[step] for step in steps))

        for local_index, local in enumerate(locais, start=1):
            log(f"[{local_index}/{len(locais)}] Unidade {local}")
            try:
                page.goto(f"{dashboard_url}/")
                page.wait_for_timeout(4000)
                selecionar_local(page, local, log)
                abrir_menu_produtos(page, log)
            except Exception as exc:
                message = str(exc)
                log(f"ERRO -> {local}: {message}")
                for product in product_rows:
                    rows.append(
                        {
                            "unidade": local,
                            "sku": product["sku"],
                            "produto": product.get("produto", ""),
                            "preco_planilha_centavos": product["preco_centavos"],
                            "status": "ERRO_UNIDADE",
                            "mensagem": message,
                        }
                    )
                continue

            for product_index, product in enumerate(product_rows, start=1):
                sku = str(product["sku"])
                source_info = fiscal_cache.get(sku, {})
                preco_origem = source_info.get("preco_centavos")
                preco_informado = product.get("preco_centavos")
                preco_resolvido = preco_origem if preco_origem is not None else preco_informado
                try:
                    if preco_resolvido in (None, ""):
                        raise RuntimeError(
                            source_errors.get(sku)
                            or "Preco nao encontrado na origem e nao foi informado na entrada"
                        )
                    preco_centavos = int(preco_resolvido)
                    product_for_update = {**product, "preco_centavos": preco_centavos}
                    log(
                        f"  [{product_index}/{len(product_rows)}] SKU #{sku} -> "
                        f"{preco_centavos} centavos"
                        + (" (origem)" if preco_origem is not None else " (entrada)")
                    )
                    buscar_produto(page, sku, log)
                    table_info = produto_na_tabela(page, sku)
                    if not table_info.get("found"):
                        if "register_product" in steps:
                            try:
                                cadastrar_produto_unico(
                                    page,
                                    dashboard_url,
                                    fiscal_source,
                                    local,
                                    product_for_update,
                                    fiscal_cache,
                                    fiscal_group,
                                    log,
                                )
                                page.goto(f"{dashboard_url}/")
                                page.wait_for_timeout(2500)
                                selecionar_local(page, local, log)
                                abrir_menu_produtos(page, log)
                                buscar_produto(page, sku, log)
                                table_info = produto_na_tabela(page, sku)
                            except Exception as register_exc:
                                rows.append(
                                    {
                                        "unidade": local,
                                        "sku": sku,
                                        "produto": product.get("produto", ""),
                                        "preco_planilha_centavos": preco_centavos,
                                        "preco_origem_centavos": preco_origem if preco_origem is not None else "",
                                        "status": "ERRO_CADASTRO",
                                        "mensagem": str(register_exc),
                                    }
                                )
                                continue

                        if not table_info.get("found"):
                            rows.append(
                                {
                                    "unidade": local,
                                    "sku": sku,
                                    "produto": product.get("produto", ""),
                                    "preco_planilha_centavos": preco_centavos,
                                    "preco_origem_centavos": preco_origem if preco_origem is not None else "",
                                    "status": "NAO_ENCONTRADO",
                                    "mensagem": "Produto nao apareceu na tabela",
                                }
                            )
                            continue

                    mountable_text = normalize_key(table_info.get("produto_montavel", ""))
                    fiscal_text = table_info.get("dados_fiscais", "")
                    current_price = parse_table_price(table_info.get("preco_atual_texto"))
                    if "SIM" in mountable_text:
                        rows.append(
                            {
                                "unidade": local,
                                "sku": sku,
                                "produto": table_info.get("produto") or product.get("produto", ""),
                                "preco_planilha_centavos": preco_centavos,
                                "preco_origem_centavos": preco_origem if preco_origem is not None else "",
                                "preco_atual_centavos": current_price or "",
                                "fiscal_atual": fiscal_text,
                                "fiscal_status": "PULADO",
                                "fiscal_origem": fiscal_source,
                                "status": "EH_MONTAVEL",
                                "mensagem": "Produto e montavel; pulado",
                            }
                        )
                        continue

                    alterar_preco_inline_na_tabela(page, sku, preco_centavos, log)
                    fiscal_status = "JA_ESTAVA_OK"
                    fiscal_message = "Fiscal ja estava preenchido"
                    final_status = "ATUALIZADO_PRECO_REPLICADO" if preco_origem is not None else "ATUALIZADO"
                    final_message = (
                        f"Preco replicado de {fiscal_source}"
                        if preco_origem is not None
                        else "Preco atualizado na tabela de produtos"
                    )
                    if "adjust_fiscal" in steps:
                        if not fiscal_source:
                            fiscal_status = "SEM_ORIGEM"
                            fiscal_message = "Selecione a origem para ajustar fiscal"
                            final_status = "PRECO_ATUALIZADO_FISCAL_PENDENTE"
                            final_message = "Preco atualizado; fiscal nao ajustado por falta de origem"
                        else:
                            try:
                                subir_fiscal_produto_unico(
                                    page,
                                    dashboard_url,
                                    fiscal_source,
                                    local,
                                    product_for_update,
                                    fiscal_cache,
                                    fiscal_group,
                                    log,
                                )
                                fiscal_status = "IMPORTADO"
                                fiscal_message = f"Fiscal replicado de {fiscal_source}"
                                final_status = "ATUALIZADO_COM_FISCAL"
                                final_message = "Preco e fiscal replicados da origem"
                                page.goto(f"{dashboard_url}/")
                                page.wait_for_timeout(2500)
                                selecionar_local(page, local, log)
                                abrir_menu_produtos(page, log)
                                buscar_produto(page, sku, log)
                                table_info = produto_na_tabela(page, sku)
                            except Exception as fiscal_exc:
                                fiscal_status = "ERRO"
                                fiscal_message = str(fiscal_exc)
                                log(f"  ERRO FISCAL -> SKU #{sku}: {fiscal_message}")
                                final_status = "PRECO_ATUALIZADO_FISCAL_ERRO"
                                final_message = f"Preco atualizado; fiscal nao importado: {fiscal_message}"
                    elif fiscal_pendente(fiscal_text):
                        fiscal_status = "PULADO"
                        fiscal_message = "Fase Ajustar fiscal desmarcada"
                        final_status = "PRECO_ATUALIZADO_FISCAL_PENDENTE"
                        final_message = "Preco atualizado; ajuste fiscal pulado"

                    if "activate_product" in steps:
                        try:
                            ativar_produto_unico_na_tabela(page, sku, log)
                        except Exception as activate_exc:
                            log(f"  ERRO LIBERACAO -> SKU #{sku}: {activate_exc}")
                            final_status = "ERRO_LIBERACAO"
                            final_message = f"{final_message}; liberacao falhou: {activate_exc}"

                    rows.append(
                        {
                            "unidade": local,
                            "sku": sku,
                            "produto": table_info.get("produto") or product.get("produto", ""),
                            "preco_planilha_centavos": preco_centavos,
                            "preco_origem_centavos": preco_origem if preco_origem is not None else "",
                            "preco_atual_centavos": current_price or "",
                            "fiscal_atual": fiscal_text,
                            "fiscal_status": fiscal_status,
                            "fiscal_origem": fiscal_source,
                            "status": final_status,
                            "mensagem": final_message,
                        }
                    )
                except Exception as exc:
                    message = str(exc)
                    log(f"  ERRO -> SKU #{sku}: {message}")
                    rows.append(
                        {
                            "unidade": local,
                            "sku": sku,
                            "produto": product.get("produto", ""),
                            "preco_planilha_centavos": product.get("preco_centavos", ""),
                            "preco_origem_centavos": preco_origem if preco_origem is not None else "",
                            "status": "ERRO",
                            "mensagem": message,
                        }
                    )
                    try:
                        page.screenshot(path=f"erro_produto_unico_{local_index}_{product_index}.png", full_page=True)
                    except Exception:
                        pass

        time.sleep(2)
        browser.close()

    report_path = salvar_relatorio_produto_unico(rows)
    log(f"Relatorio gerado: {report_path}")
    return {"rows": rows, "report_path": str(report_path), "report_filename": report_path.name}
