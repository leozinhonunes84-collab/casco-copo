from __future__ import annotations

import os
import re
import time
from base64 import b64decode
from csv import reader as csv_reader
from io import BytesIO, StringIO
from collections.abc import Callable
from datetime import datetime
from pathlib import Path

import openpyxl
from playwright.sync_api import Page, sync_playwright

from ajuste_precos_dashboard import (
    abrir_edicao,
    abrir_menu_produtos,
    bool_env,
    buscar_produto,
    env_required,
    fechar_modal,
    salvar,
    selecionar_local,
)
from sheets_prices import load_env_file


LogFn = Callable[[str], None]


def log_default(message: str) -> None:
    print(message)


def normalize_code(value: object) -> str:
    return re.sub(r"\D+", "", str(value or "").strip())


def parse_product_codes(raw_codes: object) -> list[str]:
    if isinstance(raw_codes, list):
        parts = [str(item) for item in raw_codes]
    else:
        parts = re.split(r"[\s,;]+", str(raw_codes or ""))

    codes: list[str] = []
    seen: set[str] = set()
    for part in parts:
        code = normalize_code(part)
        if code and code not in seen:
            codes.append(code)
            seen.add(code)
    if not codes:
        raise ValueError("Informe pelo menos um codigo de produto para flagar")
    return codes


def _sku_from_cell(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if text.startswith("#"):
        text = text[1:]
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    return normalize_code(text)


def _header_index(headers: list[object]) -> int:
    wanted = {"SKU", "CODIGO", "CÓDIGO", "CODIGO PRODUTO", "CÓDIGO PRODUTO", "ID", "PRODUTO"}
    for index, header in enumerate(headers):
        header_key = re.sub(r"\s+", " ", str(header or "").strip().upper())
        if header_key in wanted or "SKU" in header_key:
            return index
    return 0


def _looks_like_header(row: list[object]) -> bool:
    return any(re.search(r"[A-Za-zÀ-ÿ]", str(cell or "")) for cell in row)


def parse_product_skus(raw_codes: object = "", file_info: dict[str, object] | None = None) -> list[str]:
    if not file_info:
        return parse_product_codes(raw_codes)

    filename = str(file_info.get("name") or "skus.xlsx").strip()
    content_b64 = str(file_info.get("content_base64") or "")
    if not content_b64:
        raise ValueError("Arquivo de SKUs vazio")

    data = b64decode(content_b64)
    suffix = Path(filename).suffix.lower()
    skus: list[str] = []

    if suffix in {".xlsx", ".xlsm"}:
        workbook = openpyxl.load_workbook(BytesIO(data), data_only=True)
        worksheet = workbook.active
        header = [worksheet.cell(1, col).value for col in range(1, worksheet.max_column + 1)]
        sku_col = _header_index(header) + 1
        start_row = 2 if _looks_like_header(header) else 1
        for row_idx in range(start_row, worksheet.max_row + 1):
            sku = _sku_from_cell(worksheet.cell(row_idx, sku_col).value)
            if sku:
                skus.append(sku)
    elif suffix == ".csv":
        text = data.decode("utf-8-sig")
        sample = text[:2000]
        delimiter = ";" if sample.count(";") >= sample.count(",") else ","
        rows = list(csv_reader(StringIO(text), delimiter=delimiter))
        if rows:
            sku_index = _header_index(rows[0])
            start = 1 if _looks_like_header(rows[0]) else 0
            for row in rows[start:]:
                if sku_index < len(row):
                    sku = _sku_from_cell(row[sku_index])
                    if sku:
                        skus.append(sku)
    else:
        raise ValueError("Arquivo de SKUs deve ser .xlsx, .xlsm ou .csv")

    unique: list[str] = []
    seen: set[str] = set()
    for sku in skus:
        if sku not in seen:
            unique.append(sku)
            seen.add(sku)
    if not unique:
        raise ValueError("Nenhum SKU encontrado na planilha")
    return unique


def marcar_flags_do_produto_montavel(page: Page, log: LogFn = log_default) -> dict[str, object]:
    page.wait_for_timeout(1200)
    flags = page.evaluate(
        """
        () => {
            const visible = (el) => !!(
                el
                && (el.offsetWidth || el.offsetHeight || el.getClientRects().length)
                && getComputedStyle(el).visibility !== "hidden"
                && getComputedStyle(el).display !== "none"
            );
            return [...document.querySelectorAll("input[id^='modification-']")]
                .filter(visible)
                .map((checkbox, index) => {
                    const rect = checkbox.getBoundingClientRect();
                    return {
                        index,
                        checked: !!checkbox.checked,
                        x: rect.left + rect.width / 2,
                        y: rect.top + rect.height / 2,
                    };
                });
        }
        """
    )

    total = len(flags)
    if total == 0:
        return {
            "status": "SEM_MONTAVEL",
            "total_itens": 0,
            "marcados": 0,
            "ja_marcados": 0,
            "mensagem": "Produto nao esta montavel ou nao tem itens montaveis",
        }

    marcados = 0
    ja_marcados = 0
    for flag in flags:
        if flag.get("checked"):
            ja_marcados += 1
            continue
        page.mouse.click(float(flag["x"]), float(flag["y"]))
        page.wait_for_timeout(250)
        marcados += 1

    if marcados:
        log(f"  {marcados}/{total} item(ns) de alteracao marcado(s)")
        status = "MARCADO"
    else:
        log(f"  Todos os {total} item(ns) ja estavam marcados")
        status = "JA_MARCADO"

    return {
        "status": status,
        "total_itens": total,
        "marcados": marcados,
        "ja_marcados": ja_marcados,
        "mensagem": "OK",
    }


def flagar_itens_alteracao(page: Page, codigos: list[str], log: LogFn = log_default) -> list[dict[str, str]]:
    page.wait_for_timeout(1500)
    state = page.evaluate(
        """
        (codigos) => {
            const normalize = (value) => String(value || "")
                .normalize("NFD")
                .replace(/[\\u0300-\\u036f]/g, "")
                .replace(/\\s+/g, " ")
                .trim()
                .toUpperCase();
            const digits = (value) => String(value || "").replace(/\\D+/g, "");
            const visible = (el) => !!(
                el
                && (el.offsetWidth || el.offsetHeight || el.getClientRects().length)
                && getComputedStyle(el).visibility !== "hidden"
                && getComputedStyle(el).display !== "none"
            );
            const collectText = (container) => {
                const chunks = [
                    container.innerText,
                    container.textContent,
                    ...[...container.querySelectorAll("[title], [aria-label], input")]
                        .map((el) => [
                            el.getAttribute("title"),
                            el.getAttribute("aria-label"),
                            el.value,
                        ].filter(Boolean).join(" "))
                ];
                return normalize(chunks.filter(Boolean).join(" "));
            };
            const itemContainer = (checkbox) => {
                let current = checkbox;
                for (let depth = 0; current && depth < 10; depth += 1, current = current.parentElement) {
                    const text = normalize(current.innerText || current.textContent || "");
                    const checkboxCount = current.querySelectorAll("input[id^='modification-']").length;
                    const rect = current.getBoundingClientRect();
                    if (
                        checkboxCount === 1
                        && text.includes("ITEM DE ALTERACAO")
                        && text.includes("ITEM")
                        && rect.height > 35
                        && rect.height < 360
                    ) {
                        return current;
                    }
                }
                return checkbox.closest("label, div") || checkbox;
            };

            const checkboxes = [...document.querySelectorAll("input[id^='modification-']")].filter(visible);
            const rows = checkboxes.map((checkbox, index) => {
                const container = itemContainer(checkbox);
                const text = collectText(container);
                const rect = checkbox.getBoundingClientRect();
                return {
                    index,
                    checked: !!checkbox.checked,
                    text,
                    digitText: digits(text),
                    x: rect.left + rect.width / 2,
                    y: rect.top + rect.height / 2,
                };
            });

            return codigos.map((codigo) => {
                const code = digits(codigo);
                const row = rows.find((candidate) => {
                    if (!code) return false;
                    const tokens = candidate.text.split(/[^0-9A-Z]+/).filter(Boolean);
                    return tokens.includes(code) || candidate.digitText.includes(code);
                });
                if (!row) {
                    return {
                        code,
                        found: false,
                        checked: false,
                        changed: false,
                        status: "NAO_APARECEU",
                        message: "Codigo nao apareceu na lista de itens",
                    };
                }
                return {
                    code,
                    found: true,
                    checked: row.checked,
                    changed: false,
                    status: row.checked ? "JA_MARCADO" : "PENDENTE",
                    message: row.text,
                    x: row.x,
                    y: row.y,
                };
            });
        }
        """,
        codigos,
    )

    output: list[dict[str, str]] = []
    for item in state:
        code = str(item.get("code") or "")
        if not item.get("found"):
            log(f"  #{code}: nao apareceu na lista")
            output.append({"code": code, "status": "NAO_APARECEU", "message": str(item.get("message") or "")})
            continue
        if item.get("checked"):
            log(f"  #{code}: ja estava marcado")
            output.append({"code": code, "status": "JA_MARCADO", "message": str(item.get("message") or "")})
            continue

        page.mouse.click(float(item["x"]), float(item["y"]))
        page.wait_for_timeout(500)
        output.append({"code": code, "status": "MARCADO", "message": str(item.get("message") or "")})
        log(f"  #{code}: marcado")

    return output


def salvar_relatorio(rows: list[dict[str, object]]) -> Path:
    destino = Path(__file__).resolve().parent / "EXPORTACOES_ZIGPAY"
    destino.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    caminho = destino / f"RELATORIO_FLAG_ALTERACAO_{timestamp}.xlsx"

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Resultado"
    headers = [
        "unidade",
        "sku",
        "produto",
        "codigo",
        "status",
        "total_itens",
        "marcados",
        "ja_marcados",
        "mensagem",
    ]
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])
    for column in worksheet.columns:
        width = max(len(str(cell.value or "")) for cell in column) + 2
        worksheet.column_dimensions[column[0].column_letter].width = min(max(width, 12), 80)
    workbook.save(caminho)
    return caminho


def run_flagar_alteracao_produtos_dashboard(
    locais: list[str],
    skus: object = "",
    log: LogFn = log_default,
    file_info: dict[str, object] | None = None,
) -> dict[str, object]:
    load_env_file()
    org = env_required("ZIG_ORG")
    user = env_required("ZIG_USER")
    password = env_required("ZIG_PASSWORD")
    dashboard_url = os.environ.get("DASHBOARD_URL", "https://dashboard.zigpay.com.br").rstrip("/")

    locais = [str(local).strip().upper() for local in locais if str(local).strip()]
    sku_list = parse_product_skus(skus, file_info)
    if not locais:
        raise ValueError("Selecione pelo menos uma unidade")

    rows: list[dict[str, object]] = []

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(
            headless=bool_env("HEADLESS", default=False),
            slow_mo=60,
        )
        page = browser.new_page(viewport={"width": 1440, "height": 900})

        log("Login no Dashboard ZigPay...")
        page.goto(f"{dashboard_url}/login")
        page.wait_for_selector("input[name='organization']", timeout=15000)
        page.fill("input[name='organization']", org)
        page.fill("input[name='username']", user)
        page.fill("input[name='password']", password)
        page.click("button[type='submit']")
        try:
            page.wait_for_url(lambda url: "/login" not in url, timeout=20000)
        except Exception:
            raise RuntimeError(
                "Login falhou: verifique usuario/senha/organizacao no arquivo .env "
                "(ZIG_USER, ZIG_PASSWORD, ZIG_ORG)"
            )

        for local_index, local in enumerate(locais, start=1):
            log(f"[{local_index}/{len(locais)}] Unidade {local}")
            try:
                page.goto(f"{dashboard_url}/")
                page.wait_for_timeout(4000)
                selecionar_local(page, local, log)
                abrir_menu_produtos(page, log)
            except Exception as exc:
                message = str(exc)
                log(f"ERRO -> {local}: {message}")
                for sku in sku_list:
                    rows.append(
                        {
                            "unidade": local,
                            "sku": sku,
                            "status": "ERRO_UNIDADE",
                            "mensagem": message,
                        }
                    )
                continue

            for sku_index, sku in enumerate(sku_list, start=1):
                log(f"  [{sku_index}/{len(sku_list)}] SKU #{sku}")
                try:
                    buscar_produto(page, sku, log)
                    abrir_edicao(page, log, sku)
                except Exception as exc:
                    message = str(exc)
                    log(f"  SKU #{sku}: produto nao encontrado ({message})")
                    rows.append(
                        {
                            "unidade": local,
                            "sku": sku,
                            "status": "NAO_ENCONTRADO",
                            "mensagem": message,
                        }
                    )
                    try:
                        page.keyboard.press("Escape")
                        page.wait_for_timeout(700)
                    except Exception:
                        pass
                    continue

                try:
                    resultado = marcar_flags_do_produto_montavel(page, log)
                    if resultado["status"] == "MARCADO":
                        salvar(page, log)
                    fechar_modal(page, log)
                    rows.append(
                        {
                            "unidade": local,
                            "sku": sku,
                            **resultado,
                        }
                    )
                except Exception as exc:
                    message = str(exc)
                    log(f"  ERRO -> SKU #{sku}: {message}")
                    rows.append(
                        {
                            "unidade": local,
                            "sku": sku,
                            "status": "ERRO",
                            "mensagem": message,
                        }
                    )
                    try:
                        page.screenshot(path=f"erro_flag_alteracao_{local_index}_{sku_index}.png", full_page=True)
                    except Exception:
                        pass
                    try:
                        fechar_modal(page, log)
                    except Exception:
                        pass

        time.sleep(2)
        browser.close()

    report_path = salvar_relatorio(rows)
    log(f"Relatorio gerado: {report_path}")
    return {"rows": rows, "report_path": str(report_path), "report_filename": report_path.name}


def run_flagar_alteracao_dashboard(
    produto: str,
    locais: list[str],
    codigos: object,
    log: LogFn = log_default,
    sku_principal: str | None = None,
) -> dict[str, object]:
    load_env_file()
    org = env_required("ZIG_ORG")
    user = env_required("ZIG_USER")
    password = env_required("ZIG_PASSWORD")
    dashboard_url = os.environ.get("DASHBOARD_URL", "https://dashboard.zigpay.com.br").rstrip("/")

    produto = str(produto or "").strip().upper()
    locais = [str(local).strip().upper() for local in locais if str(local).strip()]
    codigos_lista = parse_product_codes(codigos)
    sku_principal = normalize_code(sku_principal)

    if not produto and not sku_principal:
        raise ValueError("Informe o produto principal ou o SKU principal")
    if not locais:
        raise ValueError("Selecione pelo menos uma unidade")

    rows: list[dict[str, str]] = []

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(
            headless=bool_env("HEADLESS", default=False),
            slow_mo=60,
        )
        page = browser.new_page(viewport={"width": 1440, "height": 900})

        log("Login no Dashboard ZigPay...")
        page.goto(f"{dashboard_url}/login")
        page.wait_for_selector("input[name='organization']", timeout=15000)
        page.fill("input[name='organization']", org)
        page.fill("input[name='username']", user)
        page.fill("input[name='password']", password)
        page.click("button[type='submit']")
        try:
            page.wait_for_url(lambda url: "/login" not in url, timeout=20000)
        except Exception:
            raise RuntimeError(
                "Login falhou: verifique usuario/senha/organizacao no arquivo .env "
                "(ZIG_USER, ZIG_PASSWORD, ZIG_ORG)"
            )

        for index, local in enumerate(locais, start=1):
            log(f"[{index}/{len(locais)}] Verificando flags em {local}")
            try:
                page.goto(f"{dashboard_url}/")
                page.wait_for_timeout(4000)
                selecionar_local(page, local, log)
                abrir_menu_produtos(page, log)
                buscar_produto(page, sku_principal or produto, log)
                abrir_edicao(page, log, sku_principal or None)

                resultados = flagar_itens_alteracao(page, codigos_lista, log)
                mudou = any(item["status"] == "MARCADO" for item in resultados)
                if mudou:
                    salvar(page, log)
                else:
                    log("Nenhuma alteracao nova para salvar.")
                fechar_modal(page, log)

                for item in resultados:
                    rows.append(
                        {
                            "unidade": local,
                            "produto": produto or f"SKU {sku_principal}",
                            "codigo": item["code"],
                            "status": item["status"],
                            "mensagem": item.get("message", ""),
                        }
                    )
            except Exception as exc:
                message = str(exc)
                log(f"ERRO -> {local}: {message}")
                for code in codigos_lista:
                    rows.append(
                        {
                            "unidade": local,
                            "produto": produto or f"SKU {sku_principal}",
                            "codigo": code,
                            "status": "ERRO",
                            "mensagem": message,
                        }
                    )
                try:
                    page.screenshot(path=f"erro_flag_alteracao_{index}.png", full_page=True)
                except Exception:
                    pass
                try:
                    fechar_modal(page, log)
                except Exception:
                    pass

        time.sleep(2)
        browser.close()

    report_path = salvar_relatorio(rows)
    log(f"Relatorio gerado: {report_path}")
    return {"rows": rows, "report_path": str(report_path)}
