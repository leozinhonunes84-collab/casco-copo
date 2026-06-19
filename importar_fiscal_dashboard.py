from __future__ import annotations

import os
import re
import time
import unicodedata
from collections.abc import Callable
from pathlib import Path

import openpyxl
from playwright.sync_api import Locator, Page, sync_playwright

from ajuste_precos_dashboard import abrir_menu_produtos, bool_env, buscar_produto, env_required, selecionar_local
from sheets_prices import load_env_file


LogFn = Callable[[str], None]


def log_default(message: str) -> None:
    print(message)


def normalize_key(value: object) -> str:
    text = str(value or "").strip()
    text = "".join(
        char
        for char in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(char)
    )
    return re.sub(r"\s+", " ", text).upper()


def safe_filename(value: str) -> str:
    text = normalize_key(value)
    text = re.sub(r"[^A-Z0-9]+", "_", text).strip("_")
    return text or "ARQUIVO"


def click_by_text(page: Page, text: str, log: LogFn = log_default, timeout: int = 12000) -> None:
    locators = [
        page.get_by_role("button", name=text).first,
        page.get_by_text(text, exact=True).first,
        page.locator(f"xpath=//*[normalize-space(.)='{text}']").first,
    ]
    last_error: Exception | None = None
    for locator in locators:
        try:
            locator.wait_for(state="visible", timeout=timeout)
            locator.scroll_into_view_if_needed(timeout=timeout)
            locator.click(timeout=timeout)
            return
        except Exception as exc:
            last_error = exc
    raise RuntimeError(f"Elemento nao encontrado/clicavel: {text}") from last_error


def first_enabled_button_by_text(page: Page, texts: list[str], timeout: int = 30000) -> Locator | None:
    wanted = [normalize_key(text) for text in texts]
    end_time = time.monotonic() + timeout / 1000
    while time.monotonic() < end_time:
        buttons = page.locator("button")
        count = buttons.count()
        for index in range(count):
            candidate = buttons.nth(index)
            try:
                label = normalize_key(candidate.inner_text(timeout=500))
                if not any(text in label for text in wanted):
                    continue
                if candidate.is_visible(timeout=500) and candidate.is_enabled(timeout=500):
                    return candidate
            except Exception:
                pass
        page.wait_for_timeout(500)
    return None


def click_dropdown_item(
    page: Page,
    trigger_text: str,
    item_text: str,
    log: LogFn = log_default,
    attempts: int = 4,
) -> None:
    trigger = page.get_by_role("button", name=trigger_text, exact=True).first
    trigger.wait_for(state="visible", timeout=15000)

    item_locators = [
        page.get_by_role("menuitem", name=item_text, exact=True).first,
        page.locator("li[role='menuitem']").filter(has_text=item_text).first,
        page.get_by_text(item_text, exact=True).first,
    ]
    last_error: Exception | None = None

    for attempt in range(1, attempts + 1):
        try:
            page.keyboard.press("Escape")
            trigger.scroll_into_view_if_needed(timeout=3000)
            trigger.hover(timeout=3000)
            trigger.click(timeout=3000)

            for item in item_locators:
                try:
                    item.wait_for(state="visible", timeout=1500)
                    item.click(timeout=3000)
                    return
                except Exception as exc:
                    last_error = exc
        except Exception as exc:
            last_error = exc

        if attempt < attempts:
            log(f"Dropdown nao abriu; tentando novamente ({attempt + 1}/{attempts})...")
            page.wait_for_timeout(500)

    raise RuntimeError(f"Elemento nao encontrado/clicavel: {item_text}") from last_error


def abrir_modal_importacao(page: Page, log: LogFn = log_default) -> None:
    log("Abrindo importacao em massa...")
    click_dropdown_item(page, "Adicionar Produto", "Importar produtos em massa", log)
    page.wait_for_timeout(2000)
    page.locator("input[type='file']").wait_for(state="attached", timeout=15000)


def abrir_modal_edicao_produtos_excel(page: Page, log: LogFn = log_default) -> None:
    log("Abrindo edicao de produtos via excel...")
    click_by_text(page, "Editar produtos via excel", log)
    page.wait_for_timeout(2000)
    page.locator("input[type='file']").wait_for(state="attached", timeout=15000)


def desativar_status_ativo_no_modal(page: Page, log: LogFn = log_default) -> None:
    finder_script = """
    () => {
        const marker = "data-zigpay-import-status-active-switch";
        const normalize = (value) => (value || "")
            .normalize("NFD")
            .replace(/[\\u0300-\\u036f]/g, "")
            .replace(/\\s+/g, " ")
            .trim()
            .toUpperCase();
        const targetText = "IMPORTAR PRODUTOS COM O STATUS ATIVO";
        const visible = (el) => !!(
            el
            && (el.offsetWidth || el.offsetHeight || el.getClientRects().length)
            && getComputedStyle(el).visibility !== "hidden"
            && getComputedStyle(el).display !== "none"
        );
        const checked = (switchEl) => switchEl.getAttribute("aria-checked") === "true"
            || switchEl.classList.contains("ant-switch-checked");
        document.querySelectorAll(`[${marker}]`).forEach((el) => el.removeAttribute(marker));
        const switchSelector = "button[role='switch'], .ant-switch";
        const switches = [...document.querySelectorAll(switchSelector)].filter(visible);
        const distanceFrom = (switchEl, textEl) => {
            const switchRect = switchEl.getBoundingClientRect();
            const textRect = textEl.getBoundingClientRect();
            const switchY = (switchRect.top + switchRect.bottom) / 2;
            const textY = (textRect.top + textRect.bottom) / 2;
            const horizontalGap = Math.max(0, textRect.left - switchRect.right);
            return Math.abs(switchY - textY) + horizontalGap;
        };
        const textElements = [...document.querySelectorAll("label, span, p, div")]
            .filter((el) => {
                if (!visible(el)) return false;
                const text = normalize(el.innerText);
                if (!text.includes(targetText)) return false;
                const rect = el.getBoundingClientRect();
                return text === targetText || (text.length <= targetText.length + 20 && rect.width < 600);
            })
            .sort((a, b) => {
                const aExact = normalize(a.innerText) === targetText ? 0 : 1;
                const bExact = normalize(b.innerText) === targetText ? 0 : 1;
                if (aExact !== bExact) return aExact - bExact;
                return normalize(a.innerText).length - normalize(b.innerText).length;
            });

        let target = null;
        for (const textEl of textElements) {
            let ancestor = textEl.parentElement;
            for (let depth = 0; ancestor && depth < 8 && !target; depth += 1, ancestor = ancestor.parentElement) {
                const nearby = [...ancestor.querySelectorAll(switchSelector)]
                    .filter(visible)
                    .map((switchEl) => ({ switchEl, distance: distanceFrom(switchEl, textEl) }))
                    .sort((a, b) => a.distance - b.distance);
                if (nearby.length && nearby[0].distance < 90) {
                    target = nearby[0].switchEl;
                }
            }
            if (target) break;

            const sameLine = switches
                .map((switchEl) => ({ switchEl, distance: distanceFrom(switchEl, textEl) }))
                .sort((a, b) => a.distance - b.distance);
            if (sameLine.length && sameLine[0].distance < 90) {
                target = sameLine[0].switchEl;
                break;
            }
        }

        if (!target) {
            const containers = [...document.querySelectorAll(
                "[role='dialog'], .ant-modal, .ant-modal-content, .ant-drawer, .ant-drawer-content, .ant-drawer-content-wrapper, aside, section"
            )]
                .filter((el) => {
                    const text = normalize(el.innerText);
                    return visible(el) && text.includes("IMPORTAR PRODUTOS") && text.includes(targetText);
                })
                .sort((a, b) => {
                    const ar = a.getBoundingClientRect();
                    const br = b.getBoundingClientRect();
                    return (ar.width * ar.height) - (br.width * br.height);
                });
            for (const container of containers) {
                const containerSwitches = [...container.querySelectorAll("button[role='switch'], .ant-switch")]
                    .filter(visible);
                if (containerSwitches.length === 1) {
                    target = containerSwitches[0];
                    break;
                }
            }
        }

        if (!target) return { found: false };
        target.setAttribute(marker, "true");
        const rect = target.getBoundingClientRect();
        return {
            found: true,
            checked: checked(target),
            classes: target.className || "",
            ariaChecked: target.getAttribute("aria-checked") || "",
            title: target.getAttribute("title") || "",
            x: rect.left + rect.width / 2,
            y: rect.top + rect.height / 2,
            width: rect.width,
            height: rect.height,
        };
    }
    """
    try:
        state = page.evaluate(finder_script)
        if not state.get("found"):
            raise RuntimeError(
                "Opcao 'Importar produtos com o status ativo' nao localizada; "
                "importacao cancelada para evitar produto ativo."
            )

        log(
            "Status ativo localizado: "
            f"marcado={state.get('checked')} "
            f"classe={state.get('classes')!r} "
            f"aria={state.get('ariaChecked')!r} "
            f"titulo={state.get('title')!r}"
        )

        if state.get("checked"):
            switch = page.locator("[data-zigpay-import-status-active-switch='true']").first
            try:
                switch.click(timeout=5000, force=True)
            except Exception:
                page.mouse.click(float(state["x"]), float(state["y"]))
            page.wait_for_timeout(1200)
            confirmar = first_enabled_button_by_text(
                page,
                ["Confirmar", "Sim", "OK", "Desativar", "Continuar"],
                timeout=2500,
            )
            if confirmar:
                confirmar.click()
                page.wait_for_timeout(1200)

        state = page.evaluate(finder_script)
        if state.get("found") and state.get("checked"):
            page.evaluate(
                """
                () => {
                    const target = document.querySelector("[data-zigpay-import-status-active-switch='true']");
                    if (!target) return false;
                    target.click();
                    return true;
                }
                """
            )
            page.wait_for_timeout(1200)
            confirmar = first_enabled_button_by_text(
                page,
                ["Confirmar", "Sim", "OK", "Desativar", "Continuar"],
                timeout=2500,
            )
            if confirmar:
                confirmar.click()
                page.wait_for_timeout(1200)

        for _ in range(5):
            state = page.evaluate(finder_script)
            if state.get("found") and not state.get("checked"):
                log("Opcao 'Importar produtos com o status ativo' desativada.")
                return
            page.wait_for_timeout(700)

        try:
            debug_path = Path(__file__).resolve().parent / "debug_status_ativo_modal.png"
            page.screenshot(path=str(debug_path), full_page=True)
        except Exception:
            pass

        raise RuntimeError(
            "Opcao 'Importar produtos com o status ativo' continuou marcada; "
            "importacao cancelada para evitar produto ativo. "
            f"Estado final: classe={state.get('classes')!r}, "
            f"aria={state.get('ariaChecked')!r}, titulo={state.get('title')!r}, "
            f"x={state.get('x')}, y={state.get('y')}, "
            f"w={state.get('width')}, h={state.get('height')}."
        )
    except Exception as exc:
        log(f"Nao foi possivel ajustar o status ativo no modal: {exc}")
        raise


def header_column(worksheet, header: str) -> int:
    wanted = normalize_key(header)
    for col in range(1, worksheet.max_column + 1):
        if normalize_key(worksheet.cell(1, col).value) == wanted:
            return col
    raise RuntimeError(f"Cabecalho nao encontrado no modelo ZigPay: {header}")


def sheet_value(workbook, sheet_name: str, wanted: str) -> object:
    worksheet = workbook[sheet_name]
    wanted_key = normalize_key(wanted)
    for row in range(1, worksheet.max_row + 1):
        value = worksheet.cell(row, 1).value
        if normalize_key(value) == wanted_key:
            return value
    raise RuntimeError(f"Valor nao existe no modelo ZigPay ({sheet_name}): {wanted}")


def read_first_data_row(path: Path) -> dict[str, object]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    worksheet = workbook.active
    row: dict[str, object] = {}
    for col in range(1, worksheet.max_column + 1):
        header = normalize_key(worksheet.cell(1, col).value)
        if header:
            row[header] = worksheet.cell(2, col).value
    return row


def read_data_rows(path: Path) -> list[dict[str, object]]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    worksheet = workbook.active
    headers = [
        normalize_key(worksheet.cell(1, col).value)
        for col in range(1, worksheet.max_column + 1)
    ]
    rows: list[dict[str, object]] = []
    for row_idx in range(2, worksheet.max_row + 1):
        row: dict[str, object] = {}
        has_value = False
        for col, header in enumerate(headers, start=1):
            if not header:
                continue
            value = worksheet.cell(row_idx, col).value
            if value not in (None, ""):
                has_value = True
            row[header] = value
        if has_value:
            rows.append(row)
    return rows


def produto_da_planilha(path: Path) -> str:
    produto = read_first_data_row(path).get("NOME *")
    if not produto:
        raise RuntimeError(f"Produto nao encontrado na planilha fiscal: {path}")
    return str(produto).strip()


def produtos_da_planilha(path: Path) -> list[str]:
    produtos = [str(row.get("NOME *") or "").strip() for row in read_data_rows(path)]
    produtos = [produto for produto in produtos if produto]
    if not produtos:
        raise RuntimeError(f"Produto nao encontrado na planilha fiscal: {path}")
    return produtos


def parse_price(value: str) -> float:
    match = re.search(r"R\$\s*([\d.,]+)", value)
    if not match:
        return 0
    text = match.group(1).replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0


def localizar_produto_existente(
    page: Page,
    produto: str,
    log: LogFn = log_default,
    sku_esperado: str | None = None,
) -> dict[str, object]:
    produto_key = normalize_key(produto)
    sku_key = str(sku_esperado or "").strip().removeprefix("#")
    rows = page.locator("table tbody tr")
    candidatos: list[dict[str, object]] = []

    for index in range(rows.count()):
        row = rows.nth(index)
        try:
            cells = [cell.strip() for cell in row.locator("td").all_inner_texts()]
        except Exception:
            continue

        nome_exato = any(normalize_key(cell) == produto_key for cell in cells)
        sku_exato = any(cell.strip().removeprefix("#") == sku_key for cell in cells) if sku_key else False
        if not (sku_exato if sku_key else nome_exato):
            continue

        sku = ""
        price = 0.0
        for cell in cells:
            value = cell.strip()
            if value.startswith("#"):
                sku = value[1:].strip()
            if value.startswith("R$"):
                price = parse_price(value)
        product_id = row.get_attribute("data-row-key") or ""
        if sku or product_id:
            candidatos.append({"sku": sku, "id": product_id, "price": price})

    numericos = [item for item in candidatos if re.fullmatch(r"\d+", str(item.get("sku") or ""))]
    if numericos:
        log(f"Produto existente encontrado: SKU {numericos[0]['sku']} / ID {numericos[0]['id']}")
        return numericos[0]
    if candidatos:
        log(f"Produto existente encontrado: SKU {candidatos[0]['sku']} / ID {candidatos[0]['id']}")
        return candidatos[0]

    raise RuntimeError(f"Produto nao encontrado na unidade: {produto}")


def localizar_produtos_existentes(
    page: Page,
    produtos: list[str],
    log: LogFn = log_default,
    skus_por_produto: dict[str, str] | None = None,
) -> dict[str, dict[str, object]]:
    encontrados: dict[str, dict[str, object]] = {}
    for produto in produtos:
        produto_key = normalize_key(produto)
        encontrados[produto_key] = localizar_produto_existente(
            page,
            produto,
            log,
            (skus_por_produto or {}).get(produto_key),
        )
    log(f"Produtos localizados para ajuste fiscal: {len(encontrados)}")
    return encontrados


def selecionar_produtos_na_tabela(
    page: Page,
    produtos: list[str],
    log: LogFn = log_default,
    skus_por_produto: dict[str, str] | None = None,
) -> None:
    def estado_checkbox(produto: str, sku: str | None = None) -> dict[str, object]:
        return page.evaluate(
            """
            ({produto, sku}) => {
                const normalize = (value) => (value || "")
                    .normalize("NFD")
                    .replace(/[\\u0300-\\u036f]/g, "")
                    .replace(/\\s+/g, " ")
                    .trim()
                    .toUpperCase();
                const produtoKey = normalize(produto);
                const skuKey = String(sku || "").replace(/^#/, "").trim();
                const visible = (el) => !!(
                    el
                    && (el.offsetWidth || el.offsetHeight || el.getClientRects().length)
                    && getComputedStyle(el).visibility !== "hidden"
                    && getComputedStyle(el).display !== "none"
                );
                const checked = (el) => {
                    const input = el.matches("input") ? el : el.querySelector("input[type='checkbox']");
                    if (input) return !!input.checked;
                    if (el.getAttribute("aria-checked") === "true") return true;
                    if (el.getAttribute("aria-checked") === "false") return false;
                    return String(el.className).includes("checked");
                };
                const rows = [...document.querySelectorAll("table tbody tr")].filter(visible);
                for (const row of rows) {
                    const cells = [...row.querySelectorAll("td")].map((cell) => normalize(cell.innerText));
                    const rawCells = [...row.querySelectorAll("td")].map((cell) => cell.innerText.trim().replace(/^#/, ""));
                    if (skuKey ? !rawCells.some((cell) => cell === skuKey) : !cells.some((cell) => cell === produtoKey)) continue;

                    const checkbox = [
                        ...row.querySelectorAll("input[type='checkbox'], [role='checkbox'], .ant-checkbox, label")
                    ].filter(visible)[0];
                    if (!checkbox) {
                        return { ok: false, reason: `Checkbox nao encontrado para produto: ${produto}` };
                    }
                    const target = checkbox.closest?.("label, [role='checkbox'], .ant-checkbox-wrapper, .ant-checkbox") || checkbox;
                    const box = target.getBoundingClientRect();
                    return {
                        ok: true,
                        checked: checked(target),
                        x: box.left + box.width / 2,
                        y: box.top + box.height / 2,
                    };
                }
                return { ok: false, reason: `Produto nao encontrado na tabela para marcar: ${produto}` };
            }
            """,
            {"produto": produto, "sku": sku or ""},
        )

    marcados = []
    for produto in produtos:
        sku = (skus_por_produto or {}).get(normalize_key(produto))
        estado = estado_checkbox(produto, sku)
        if not estado.get("ok"):
            raise RuntimeError(str(estado.get("reason") or f"Nao foi possivel marcar: {produto}"))

        for _ in range(3):
            if estado.get("checked"):
                break
            page.mouse.click(float(estado["x"]), float(estado["y"]))
            page.wait_for_timeout(450)
            estado = estado_checkbox(produto, sku)

        if not estado.get("checked"):
            raise RuntimeError(f"Produto nao ficou marcado na tabela: {produto}")
        marcados.append(produto)

    log(f"Produtos marcados na tabela: {len(marcados)}")


def desativar_produtos_na_tabela(
    page: Page,
    produtos: list[str],
    log: LogFn = log_default,
    skus_por_produto: dict[str, str] | None = None,
) -> None:
    def confirmar_desativacao() -> None:
        confirmar = first_enabled_button_by_text(page, ["OK"], timeout=2500)
        if confirmar:
            confirmar.click()
            page.wait_for_timeout(800)

    def estado_switch_ativo(produto: str, sku: str | None = None) -> dict[str, object]:
        return page.evaluate(
            """
            ({produto, sku}) => {
                const normalize = (value) => (value || "")
                    .normalize("NFD")
                    .replace(/[\\u0300-\\u036f]/g, "")
                    .replace(/\\s+/g, " ")
                    .trim()
                    .toUpperCase();
                const produtoKey = normalize(produto);
                const skuKey = String(sku || "").replace(/^#/, "").trim();
                const visible = (el) => !!(
                    el
                    && (el.offsetWidth || el.offsetHeight || el.getClientRects().length)
                    && getComputedStyle(el).visibility !== "hidden"
                    && getComputedStyle(el).display !== "none"
                );
                const checked = (el) => {
                    const input = el.matches("input") ? el : el.querySelector("input");
                    if (input) return !!input.checked;
                    if (el.getAttribute("aria-checked") === "true") return true;
                    if (el.getAttribute("aria-checked") === "false") return false;
                    return String(el.className).includes("checked");
                };
                const rows = [...document.querySelectorAll("table tbody tr")].filter(visible);
                for (const row of rows) {
                    const cells = [...row.querySelectorAll("td")].map((cell) => normalize(cell.innerText));
                    const rawCells = [...row.querySelectorAll("td")].map((cell) => cell.innerText.trim().replace(/^#/, ""));
                    if (skuKey ? !rawCells.some((cell) => cell === skuKey) : !cells.some((cell) => cell === produtoKey)) continue;

                    const switches = [...row.querySelectorAll("button[role='switch'], .ant-switch, [role='switch']")]
                        .filter(visible);
                    if (!switches.length) {
                        return { ok: false, reason: `Switch Ativado nao encontrado para produto: ${produto}` };
                    }
                    const target = switches.sort((a, b) => {
                        const ar = a.getBoundingClientRect();
                        const br = b.getBoundingClientRect();
                        return br.left - ar.left;
                    })[0];
                    const box = target.getBoundingClientRect();
                    return {
                        ok: true,
                        checked: checked(target),
                        x: box.left + box.width / 2,
                        y: box.top + box.height / 2,
                    };
                }
                return { ok: false, reason: `Produto nao encontrado na tabela para desativar: ${produto}` };
            }
            """,
            {"produto": produto, "sku": sku or ""},
        )

    desativados = []
    for produto in produtos:
        sku = (skus_por_produto or {}).get(normalize_key(produto))
        estado = estado_switch_ativo(produto, sku)
        if not estado.get("ok"):
            raise RuntimeError(str(estado.get("reason") or f"Nao foi possivel desativar: {produto}"))

        for _ in range(3):
            if not estado.get("checked"):
                break
            page.mouse.click(float(estado["x"]), float(estado["y"]))
            page.wait_for_timeout(700)
            confirmar_desativacao()
            estado = estado_switch_ativo(produto, sku)

        if estado.get("checked"):
            raise RuntimeError(f"Produto continuou ativo apos tentar desativar: {produto}")
        desativados.append(produto)

    log(f"Produtos desativados apos fiscal: {len(desativados)}")


def baixar_tabela_exemplo(page: Page, unidade: str, log: LogFn = log_default) -> Path:
    destino = Path(__file__).resolve().parent / "TEMPLATES_ZIGPAY"
    destino.mkdir(parents=True, exist_ok=True)
    caminho = destino / f"EXEMPLO_{safe_filename(unidade)}.xlsx"

    log(f"Baixando modelo ZigPay da unidade {unidade}...")
    baixar = first_enabled_button_by_text(
        page,
        ["Baixar tabela dos produtos selecionados", "Baixar tabela exemplo"],
        timeout=30000,
    )
    if not baixar:
        raise RuntimeError(
            "Botao para baixar o modelo ZigPay nao foi encontrado "
            "(tabela exemplo ou produtos selecionados)"
        )
    label = baixar.inner_text(timeout=2000).strip()
    log(f"Modelo localizado: {label}")
    with page.expect_download(timeout=30000) as download_info:
        baixar.click(timeout=15000)
    download = download_info.value
    download.save_as(str(caminho))
    return caminho


def montar_planilha_no_modelo_zigpay(
    template_path: Path,
    fiscal_path: Path,
    unidade: str,
    produtos_existentes: dict[str, dict[str, object]],
    log: LogFn = log_default,
) -> Path:
    fiscal_rows = read_data_rows(fiscal_path)
    if not fiscal_rows:
        raise RuntimeError(f"Nenhum produto encontrado na planilha fiscal: {fiscal_path}")

    workbook = openpyxl.load_workbook(template_path)
    worksheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.active

    for row in range(2, worksheet.max_row + 1):
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row, col)
            if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                cell.value = None

    for row_idx, fiscal_row in enumerate(fiscal_rows, start=2):
        produto = fiscal_row.get("NOME *")
        if not produto:
            continue

        produto_existente = produtos_existentes.get(normalize_key(produto))
        if not produto_existente:
            raise RuntimeError(f"Produto existente nao localizado para montar fiscal: {produto}")

        grupo = fiscal_row.get("FISCAL - GRUPO FISCAL *") or "SEM ST COM PIS/COFINS NAO TRIBUTAVEIS"
        ncm = fiscal_row.get("FISCAL - NCM *") or "22030000"
        cest = fiscal_row.get("FISCAL - CEST") or "0000000"
        bares = fiscal_row.get("BARES (SEPARADOS POR PONTO E VIRGULA)") or ""

        values = {
            "NOME *": produto,
            "TIPO DE PRODUTO *": sheet_value(workbook, "Kinds", "Chopp"),
            "CATEGORIA *": sheet_value(workbook, "Categories", "CHOPE"),
            "CATEGORIA DO MENU": sheet_value(workbook, "Categories", "CHOPE"),
            "PRECO *": produto_existente.get("price") or 0,
            "BARES (SEPARADOS POR PONTO E VIRGULA)": (
                sheet_value(workbook, "Bares", "CHOPE") if normalize_key(bares) == "CHOPE" else None
            ),
            "SKU": None,
            "FISCAL - NCM *": str(ncm),
            "FISCAL - CEST": str(cest),
            "FISCAL - GRUPO FISCAL *": sheet_value(workbook, "FiscalProductGroups", str(grupo)),
            "FISCAL - PERFIL FISCAL *": sheet_value(workbook, "FiscalProfiles", unidade),
            "ESTOCAVEL": sheet_value(workbook, "Boolean", "NAO"),
            "UNIDADE DE MEDIDA": sheet_value(workbook, "Units", "Unidades"),
            "CONTEM ALCOOL?": sheet_value(workbook, "Boolean", "NAO"),
            "NAO EXIBIR PRODUTO NO APLICATIVO ZIGAPP": sheet_value(workbook, "Boolean", "NAO"),
            "ID": produto_existente.get("id"),
        }

        for header, value in values.items():
            worksheet.cell(row_idx, header_column(worksheet, header)).value = value

    destino = fiscal_path.parent / f"IMPORTAR_{fiscal_path.stem}.xlsx"
    workbook.save(destino)
    log(f"Planilha ajustada no modelo da ZigPay: {destino.name} ({len(fiscal_rows)} linhas)")
    return destino


def valor_modelo(workbook, sheet_name: str, value: object, default: str) -> object:
    text = str(value or default).strip() or default
    try:
        return sheet_value(workbook, sheet_name, text)
    except RuntimeError:
        if sheet_name == "Bares" and normalize_key(text) != normalize_key(default):
            return sheet_value(workbook, sheet_name, default)
        raise


def valor_modelo_ou_vazio(workbook, sheet_name: str, value: object, default: str) -> object:
    if not str(value or "").strip():
        return None
    return valor_modelo(workbook, sheet_name, value, default)


def montar_planilha_produtos_no_modelo_zigpay(
    template_path: Path,
    product_path: Path,
    unidade: str,
    log: LogFn = log_default,
) -> Path:
    linhas = read_data_rows(product_path)
    if not linhas:
        raise RuntimeError(f"Nenhum produto encontrado na planilha: {product_path}")

    workbook = openpyxl.load_workbook(template_path)
    worksheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.active

    for row in range(2, worksheet.max_row + 1):
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row, col)
            if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                cell.value = None

    for row_idx, linha in enumerate(linhas, start=2):
        produto = linha.get("NOME *")
        if not produto:
            continue

        values = {
            "NOME *": produto,
            "TIPO DE PRODUTO *": valor_modelo(workbook, "Kinds", linha.get("TIPO DE PRODUTO *"), "Chopp"),
            "CATEGORIA *": valor_modelo(workbook, "Categories", linha.get("CATEGORIA *"), "CHOPE"),
            "CATEGORIA DO MENU": valor_modelo(workbook, "Categories", linha.get("CATEGORIA DO MENU"), "CHOPE"),
            "PRECO *": linha.get("PRECO *") or 0,
            "PRECO EM CENTAVOS": linha.get("PRECO EM CENTAVOS") or 0,
            "BARES (SEPARADOS POR PONTO E VIRGULA)": valor_modelo_ou_vazio(
                workbook,
                "Bares",
                linha.get("BARES (SEPARADOS POR PONTO E VIRGULA)"),
                "CHOPE",
            ),
            "SKU": linha.get("SKU"),
            "FISCAL - NCM *": str(linha.get("FISCAL - NCM *") or "22030000"),
            "FISCAL - CEST": str(linha.get("FISCAL - CEST") or "0000000"),
            "FISCAL - GRUPO FISCAL *": valor_modelo(
                workbook,
                "FiscalProductGroups",
                linha.get("FISCAL - GRUPO FISCAL *"),
                "SEM ST COM PIS/COFINS NAO TRIBUTAVEIS",
            ),
            "FISCAL - PERFIL FISCAL *": valor_modelo(
                workbook,
                "FiscalProfiles",
                linha.get("FISCAL - PERFIL FISCAL *") or unidade,
                unidade,
            ),
            "ESTOCAVEL": valor_modelo(workbook, "Boolean", linha.get("ESTOCAVEL"), "NAO"),
            "UNIDADE DE MEDIDA": valor_modelo(workbook, "Units", linha.get("UNIDADE DE MEDIDA"), "Unidades"),
            "CONTEM ALCOOL?": valor_modelo(workbook, "Boolean", linha.get("CONTEM ALCOOL?"), "NAO"),
            "NAO EXIBIR PRODUTO NO APLICATIVO ZIGAPP": valor_modelo(
                workbook,
                "Boolean",
                linha.get("NAO EXIBIR PRODUTO NO APLICATIVO ZIGAPP"),
                "NAO",
            ),
            "ID": linha.get("ID"),
        }

        for header, value in values.items():
            worksheet.cell(row_idx, header_column(worksheet, header)).value = value

    destino = product_path.parent / f"IMPORTAR_{product_path.stem}.xlsx"
    workbook.save(destino)
    log(f"Planilha de produtos ajustada no modelo da ZigPay: {destino.name}")
    return destino


def extrair_contador(texto: str, label: str) -> int | None:
    match = re.search(rf"(\d+)\s+{re.escape(label)}", texto, flags=re.IGNORECASE)
    return int(match.group(1)) if match else None


def baixar_tabela_erros(page: Page, caminho_importado: Path, log: LogFn = log_default) -> Path | None:
    destino = Path(__file__).resolve().parent / "ERROS_IMPORTACAO_ZIGPAY"
    destino.mkdir(parents=True, exist_ok=True)
    caminho = destino / f"ERROS_{caminho_importado.stem}.xlsx"
    try:
        with page.expect_download(timeout=20000) as download_info:
            click_by_text(page, "Baixar tabela com erros", log, timeout=15000)
        download = download_info.value
        download.save_as(str(caminho))
        return caminho
    except Exception as exc:
        log(f"Nao foi possivel baixar a tabela com erros: {exc}")
        return None


def importar_arquivo_atual(page: Page, caminho: Path, log: LogFn = log_default) -> None:
    caminho = caminho.resolve()
    if not caminho.exists():
        raise FileNotFoundError(f"Arquivo nao encontrado: {caminho}")

    log(f"Anexando arquivo: {caminho.name}")
    page.locator("input[type='file']").set_input_files(str(caminho))
    page.wait_for_timeout(2000)

    avancar = first_enabled_button_by_text(page, ["Avan"], timeout=30000)
    if not avancar:
        raise RuntimeError("Botao Avancar nao habilitou depois de anexar o arquivo")
    log("Avancando importacao...")
    avancar.click()
    page.wait_for_timeout(6000)

    body = page.locator("body").inner_text(timeout=10000)
    texto = " ".join(body.split())
    erros = extrair_contador(texto, "Produtos com erro")
    importados = extrair_contador(texto, "Produtos importados")

    if erros and erros > 0:
        erro_path = baixar_tabela_erros(page, caminho, log)
        detalhe = f" Tabela de erros: {erro_path}" if erro_path else ""
        raise RuntimeError(f"ZigPay recusou {erros} produto(s).{detalhe}")

    finalizar = first_enabled_button_by_text(page, ["Concluir", "Finalizar"], timeout=12000)
    if finalizar:
        label = finalizar.inner_text(timeout=1000).strip()
        log(f"Finalizando importacao: {label}")
        finalizar.click()
        page.wait_for_timeout(5000)

    if importados is not None:
        log(f"Importacao aceita pela ZigPay: {importados} produto(s) importado(s).")
    else:
        log("Importacao enviada ao Dashboard.")


def run_importar_fiscal_dashboard(
    arquivos_por_unidade: list[dict[str, object]],
    log: LogFn = log_default,
) -> list[dict[str, str]]:
    load_env_file()
    org = env_required("ZIG_ORG")
    user = env_required("ZIG_USER")
    password = env_required("ZIG_PASSWORD")
    dashboard_url = os.environ.get("DASHBOARD_URL", "https://dashboard.zigpay.com.br").rstrip("/")

    if not arquivos_por_unidade:
        raise ValueError("Nenhum arquivo fiscal para importar")

    errors: list[dict[str, str]] = []

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(
            headless=bool_env("HEADLESS", default=False),
            slow_mo=60,
        )
        page = browser.new_page(viewport={"width": 1440, "height": 900})

        log("Login no Dashboard ZigPay...")
        page.goto(f"{dashboard_url}/login")
        page.wait_for_selector("input[name='organization']", timeout=15000)
        page.fill("input[name='organization']", org)
        page.fill("input[name='username']", user)
        page.fill("input[name='password']", password)
        page.click("button[type='submit']")
        try:
            page.wait_for_url(lambda url: "/login" not in url, timeout=20000)
        except Exception:
            raise RuntimeError(
                "Login falhou: verifique usuario/senha/organizacao no arquivo .env "
                "(ZIG_USER, ZIG_PASSWORD, ZIG_ORG)"
            )

        for index, item in enumerate(arquivos_por_unidade, start=1):
            unidade = str(item["target"]).strip().upper()
            caminho = Path(item["path"])
            log(f"[{index}/{len(arquivos_por_unidade)}] Importando fiscal em {unidade}")
            try:
                page.goto(f"{dashboard_url}/")
                page.wait_for_timeout(4000)
                selecionar_local(page, unidade, log)
                abrir_menu_produtos(page, log)
                produtos = produtos_da_planilha(caminho)
                precise_skus = item.get("precise_skus")
                skus_por_produto: dict[str, str] = {}
                if isinstance(precise_skus, dict):
                    sku_keys = ["MAIN", "REGUA", "P", "G", "1L"]
                    skus_por_produto = {
                        normalize_key(produto): str(precise_skus.get(key, "")).strip().removeprefix("#")
                        for produto, key in zip(produtos, sku_keys)
                    }
                log(f"Produtos na planilha fiscal: {len(produtos)}")
                buscar_produto(page, produtos[0], log)
                produtos_existentes = localizar_produtos_existentes(page, produtos, log, skus_por_produto)
                selecionar_produtos_na_tabela(page, produtos, log, skus_por_produto)
                abrir_modal_edicao_produtos_excel(page, log)
                desativar_status_ativo_no_modal(page, log)
                template_path = baixar_tabela_exemplo(page, unidade, log)
                arquivo_zigpay = montar_planilha_no_modelo_zigpay(
                    template_path,
                    caminho,
                    unidade,
                    produtos_existentes,
                    log,
                )
                importar_arquivo_atual(page, arquivo_zigpay, log)
                try:
                    page.keyboard.press("Escape")
                    page.wait_for_timeout(1000)
                except Exception:
                    pass
                buscar_produto(page, produtos[0], log)
                desativar_produtos_na_tabela(page, produtos, log, skus_por_produto)
                try:
                    page.keyboard.press("Escape")
                    page.wait_for_timeout(1000)
                except Exception:
                    pass
                log(f"OK -> Fiscal importado em {unidade}")
            except Exception as exc:
                message = str(exc)
                log(f"ERRO -> {unidade}: {message}")
                errors.append({"store": unidade, "error": message})
                try:
                    page.screenshot(path=f"erro_importar_fiscal_{index}.png", full_page=True)
                except Exception:
                    pass
                try:
                    page.keyboard.press("Escape")
                    page.wait_for_timeout(1000)
                except Exception:
                    pass

        time.sleep(2)
        browser.close()

    return errors


def run_importar_produtos_dashboard(
    arquivos_por_unidade: list[dict[str, object]],
    log: LogFn = log_default,
) -> list[dict[str, str]]:
    load_env_file()
    org = env_required("ZIG_ORG")
    user = env_required("ZIG_USER")
    password = env_required("ZIG_PASSWORD")
    dashboard_url = os.environ.get("DASHBOARD_URL", "https://dashboard.zigpay.com.br").rstrip("/")

    if not arquivos_por_unidade:
        raise ValueError("Nenhum arquivo de produtos para importar")

    errors: list[dict[str, str]] = []

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(
            headless=bool_env("HEADLESS", default=False),
            slow_mo=60,
        )
        page = browser.new_page(viewport={"width": 1440, "height": 900})

        log("Login no Dashboard ZigPay...")
        page.goto(f"{dashboard_url}/login")
        page.fill("input[name='organization']", org)
        page.fill("input[name='username']", user)
        page.fill("input[name='password']", password)
        page.keyboard.press("Enter")
        page.wait_for_timeout(7000)

        for index, item in enumerate(arquivos_por_unidade, start=1):
            unidade = str(item["target"]).strip().upper()
            caminho = Path(str(item["path"]))
            log(f"[{index}/{len(arquivos_por_unidade)}] Importando produtos em {unidade}")
            try:
                page.goto(f"{dashboard_url}/")
                page.wait_for_timeout(4000)
                selecionar_local(page, unidade, log)
                abrir_menu_produtos(page, log)
                abrir_modal_importacao(page, log)
                desativar_status_ativo_no_modal(page, log)
                template_path = baixar_tabela_exemplo(page, unidade, log)
                arquivo_zigpay = montar_planilha_produtos_no_modelo_zigpay(
                    template_path,
                    caminho,
                    unidade,
                    log,
                )
                importar_arquivo_atual(page, arquivo_zigpay, log)
                try:
                    page.keyboard.press("Escape")
                    page.wait_for_timeout(1000)
                except Exception:
                    pass

                log(f"OK -> Produtos importados em {unidade}")
            except Exception as exc:
                message = str(exc)
                log(f"ERRO -> {unidade}: {message}")
                errors.append({"store": unidade, "error": message})
                try:
                    page.screenshot(path=f"erro_importar_produtos_{index}.png", full_page=True)
                except Exception:
                    pass
                try:
                    page.keyboard.press("Escape")
                    page.wait_for_timeout(1000)
                except Exception:
                    pass

        time.sleep(2)
        browser.close()

    return errors
