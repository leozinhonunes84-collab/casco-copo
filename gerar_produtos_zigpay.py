from __future__ import annotations

import os
import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from sheets_prices import load_env_file


DEFAULT_SOURCE = r"C:\Users\ONZE COPOS\Desktop\SELENIUM\files\SUBIRPRODUTOS.xlsx"
DEFAULT_EXPORT_DIR = Path(__file__).resolve().parent / "EXPORTACOES_ZIGPAY"

LOCAIS = [
    "BREWTECO GAVEA",
    "BREWTECO TIJUCA",
    "BREWTECO LARANJEIRAS",
    "BREWTECO LEBLON",
    "BREWTECO ROSAS",
    "BREWTECO BOTAFOGO",
]

GRUPOS_FISCAIS = [
    "SEM ST COM PIS/COFINS NAO TRIBUTAVEIS",
    "COM ST PIS/COFINS TRIBUTAVEIS",
]

CABECALHO = [
    "Nome *", "Tipo de produto *", "Categoria *", "Categoria do menu",
    "ID do Produto de Sistema", "Preço *", "Preço em centavos",
    "Bares (separados por ponto e vírgula)", "Imagem (índice)",
    "Imagem (visualização)", "SKU", "Fiscal - NCM *", "Fiscal - CEST",
    "Fiscal - Grupo fiscal *", "Fiscal - Perfil fiscal *",
    "Fiscal - Base de Cálculo do ICMS Retido na operação anterior",
    "Fiscal - Valor do ICMS Próprio do Substituto",
    "Fiscal - Alíquota Suportada pelo Consumidor Final", "Estocável",
    "Unidade de medida", "Descrição", "Imagem (ID)", "Contém álcool?",
    "Não exibir produto no aplicativo ZigApp", "ID", "Imagem original",
    "ABV - Teor alcoólico", "Marca da cerveja", "Estilo da cerveja",
]


def source_path() -> Path:
    load_env_file()
    return Path(os.environ.get("PRODUCT_SOURCE_XLSX", DEFAULT_SOURCE))


def export_dir() -> Path:
    load_env_file()
    return Path(os.environ.get("PRODUCT_EXPORT_DIR", str(DEFAULT_EXPORT_DIR)))


def normalize(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().upper())


def safe_filename(value: str) -> str:
    text = normalize(value)
    text = re.sub(r"[^A-Z0-9]+", "_", text).strip("_")
    return text or "EXPORTACAO"


def montar_linha(
    nome: str,
    bares: str,
    sku: str | int,
    local: str,
    grupo: str,
    replicar_fiscal: bool = True,
) -> list[Any]:
    linha = [""] * len(CABECALHO)
    linha[0] = nome
    linha[1] = "Chopp"
    linha[2] = "CHOPE"
    linha[3] = "CHOPE"
    linha[5] = 0
    linha[6] = 0
    linha[7] = bares
    linha[10] = str(sku)
    if replicar_fiscal:
        linha[11] = "22030000"
        linha[12] = "0000000"
        linha[13] = grupo
        linha[14] = local
    linha[18] = "Não"
    linha[19] = "Unidades"
    linha[22] = "Não"
    linha[23] = "Não"
    return linha


def salvar_excel(linhas: list[list[Any]], nome_base: str, local: str) -> dict[str, str]:
    destino = export_dir()
    destino.mkdir(parents=True, exist_ok=True)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Importacao"

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=9)
    for col, titulo in enumerate(CABECALHO, start=1):
        cell = worksheet.cell(row=1, column=col, value=titulo)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 30

    for row_idx, linha in enumerate(linhas, start=2):
        for col, value in enumerate(linha, start=1):
            worksheet.cell(row=row_idx, column=col, value=value)

    worksheet.column_dimensions["A"].width = 45
    for column in ["B", "C", "D", "N", "O"]:
        worksheet.column_dimensions[column].width = 20
    worksheet.column_dimensions["K"].width = 18

    filename = f"{safe_filename(nome_base)}_{safe_filename(local)}.xlsx"
    path = destino / filename
    workbook.save(path)
    return {"path": str(path), "filename": filename}


def carregar_produtos() -> list[dict[str, str]]:
    caminho = source_path()
    if not caminho.exists():
        raise FileNotFoundError(f"SUBIRPRODUTOS.xlsx nao encontrado: {caminho}")

    workbook = openpyxl.load_workbook(caminho, data_only=True)
    worksheet = workbook.active
    produtos = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row and row[0] and len(row) > 1 and row[1]:
            produtos.append({"sku": str(row[0]).strip(), "name": normalize(row[1])})
    return produtos


def buscar_produtos(termo: str = "", limit: int = 80) -> list[dict[str, str]]:
    termo_norm = normalize(termo)
    produtos = carregar_produtos()
    if termo_norm:
        produtos = [
            produto
            for produto in produtos
            if termo_norm in produto["name"] or termo_norm in produto["sku"]
        ]
    return produtos[:limit]


def gerar_produto_novo(
    nome: str,
    sku_inicial: str,
    pular: int,
    local: str,
    grupo: str,
    replicar_fiscal: bool = True,
) -> dict[str, Any]:
    nome_norm = normalize(nome)
    if not nome_norm:
        raise ValueError("Digite o nome do chope")
    if not str(sku_inicial).strip().isdigit():
        raise ValueError("SKU inicial deve conter apenas numeros")

    sku = int(str(sku_inicial).strip()) + int(pular or 0)
    tamanhos = ["REGUA", "P", "G", "1L"]
    linhas = [montar_linha(nome_norm, "CHOPE", sku, local, grupo, replicar_fiscal)]
    for index, tamanho in enumerate(tamanhos, start=1):
        linhas.append(montar_linha(f"{nome_norm} {tamanho}", "", sku + index, local, grupo, replicar_fiscal))

    saved = salvar_excel(linhas, nome_norm, local)
    return {
        **saved,
        "mode": "new",
        "count": len(linhas),
        "target": local,
        "products": [{"name": linha[0], "sku": str(linha[10])} for linha in linhas],
        "sku_start": sku,
        "sku_end": sku + len(tamanhos),
        "next_sku": sku + len(tamanhos) + 1,
    }


def gerar_produto_existente(
    produtos: list[dict[str, str]],
    local: str,
    grupo: str,
    replicar_fiscal: bool = True,
) -> dict[str, Any]:
    if not produtos:
        raise ValueError("Selecione pelo menos 1 produto")
    if len(produtos) > 5:
        raise ValueError("Selecione no maximo 5 produtos")

    linhas = []
    for index, produto in enumerate(produtos):
        sku = str(produto.get("sku", "")).strip()
        nome = normalize(produto.get("name", ""))
        if not sku or not nome:
            raise ValueError("Produto existente sem SKU ou nome")
        linhas.append(montar_linha(nome, "CHOPE", sku, local, grupo, replicar_fiscal))

    saved = salvar_excel(linhas, produtos[0]["name"], local)
    return {
        **saved,
        "mode": "existing",
        "count": len(linhas),
        "target": local,
        "products": [{"name": linha[0], "sku": str(linha[10])} for linha in linhas],
    }


def gerar_replicacao_fiscal(
    nome: str,
    fiscal_origem: str,
    locais_destino: list[str],
    grupo: str,
) -> dict[str, Any]:
    nome_norm = normalize(nome)
    origem_norm = normalize(fiscal_origem)
    destinos_norm = [normalize(local) for local in locais_destino if normalize(local)]
    if not nome_norm:
        raise ValueError("Digite o nome do chope")
    if not origem_norm:
        raise ValueError("Selecione de onde copiar o fiscal")
    if not destinos_norm:
        raise ValueError("Marque pelo menos uma unidade de destino")

    files = []
    nome_base = f"FISCAL_{nome_norm}"
    tamanhos = ["REGUA", "P", "G", "1L"]
    for destino in destinos_norm:
        linhas = [montar_linha(nome_norm, "CHOPE", "", destino, grupo, True)]
        for tamanho in tamanhos:
            linhas.append(montar_linha(f"{nome_norm} {tamanho}", "", "", destino, grupo, True))
        saved = salvar_excel(linhas, nome_base, destino)
        files.append({
            **saved,
            "target": destino,
            "count": len(linhas),
            "products": [{"name": linha[0], "sku": str(linha[10])} for linha in linhas],
        })
    return {
        "mode": "fiscal",
        "count": sum(int(file_info["count"]) for file_info in files),
        "source": origem_norm,
        "targets": destinos_norm,
        "files": files,
        "path": files[0]["path"],
        "filename": files[0]["filename"],
    }
